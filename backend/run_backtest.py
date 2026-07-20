
from __future__ import (absolute_import, division, print_function,
                        unicode_literals)

import glob
import json
import math
import os
import shutil
import time
import traceback
from datetime import datetime, timedelta, timezone

import matplotlib
matplotlib.use('Agg')  # headless: pyfolio still renders its tear sheet to a file
import matplotlib.pyplot as plt

import backtrader as bt

from EMACrossShortTest import EMACrossShortTest
from KalmanFilter import KalmanTest
from ExtendedKalmanFilter import ExtendedKalmanTest
from AccelerationKalmanFilter import AccelerationKalmanTest
from EMAlgoTest import EMTest
from EMAlgoNonLinTest import EMNonLinTest
from EMNonLinScaled import EMNonLinScaled

import _paths
BASE = _paths.ROOT   # datas/ and reports/ live at project root
DATA_DIR = os.path.join(BASE, 'datas')
DATA_DIR_1H = os.path.join(BASE, 'datas_1h')


def _tf_cfg(timeframe):
    """(data_dir, filename_suffix, bt_compression) for a bar timeframe.
    '1m' -> minute bars in datas/; '1h' -> hourly bars in datas_1h/ loaded
    with backtrader compression=60."""
    if (timeframe or '1m') == '1h':
        return DATA_DIR_1H, '1h', 60
    return DATA_DIR, '1m', 1
REPORTS = os.path.join(BASE, 'reports')
CHARTDATA_DIR = os.path.join(REPORTS, 'chartdata')    # per-symbol interactive chart data (JSON)
TEST_DATA_DIR = os.path.join(REPORTS, 'test_data')   # per-run report archive
PYFOLIO_DIR = TEST_DATA_DIR   # back-compat alias (server.py references it)
STATUS_PATH = os.path.join(REPORTS, 'status.json')
RESULTS_PATH = os.path.join(REPORTS, 'results.json')

STARTING_CASH = 100_000.0
LEVERAGE = 10
COMMISSION = 0.0002   # per-side, as a fraction of notional (0.0002 = 0.02%)

# ---- risk metrics -------------------------------------------------------
# Bars are 1-minute and crypto trades 24/7, so a year is 365*24*60 minutes.
# Sharpe/Sortino annualize a per-period ratio by sqrt(periods per year):
#     mean(r - rf) / std(r - rf, ddof=1) * sqrt(MINUTES_PER_YEAR)
# rf is a PER-MINUTE rate (RISK_FREE_ANNUAL / MINUTES_PER_YEAR), matching how
# empyrical subtracts risk_free from every return observation.
MINUTES_PER_YEAR = 365 * 24 * 60


RISK_FREE_ANNUAL = 0.06

# ---- strategy selection -------------------------------------------------
# Swap the strategy by changing these two lines. Any PortfolioStrategy
# subclass works (see strategy_base.py); its build_chart_lines() drives the
# dashboard overlays, and STRATEGY_PARAMS is recorded in results.json.
STRATEGY = EMTest

# EMTest (EMAlgoTest.py) is the 2-state constant-velocity model with periodic
# EM refits of Q, R -- it does NOT take q_acc/k_exit/drag_m/etc (those are
# ExtendedKalmanTest-only params; passing them here raises "PortfolioStrategy
# .__init__() got an unexpected keyword argument" since they're unrecognized).
STRATEGY_PARAMS = dict(
    trade_usd=2000.0,
    k=2,             # TUNED: wider bands (k=2 over-traded, ~lost; 3 too few trades)
    warmup=180,
    reversion=False,
    q_level=0.1e-3,    # INITIAL seed only -- EM re-learns Q,R from data after warmup
    q_vel=0.1e-6,
    em_window=720,     # bars of rolling history each EM refit uses
    em_interval=1440,  # TUNED: refit rarely (frequent refits overfit Q,R to noise)
    em_iters=2,        # TUNED: few EM iters (5-10 overfit; 2 generalizes better)
)

# Registry consumed by the dashboard's strategy picker (GET /api/strategies).
# name -> (class, tuned defaults). The defaults are what the UI pre-fills and
# what a run uses for any param the request leaves out.
STRATEGIES = {
    'EMTest': (EMTest, dict(STRATEGY_PARAMS)),
    'KalmanTest': (KalmanTest, dict(
        trade_usd=2000.0, k=2.0, warmup=180, reversion=False,
        q_level=0.1e-3, q_vel=0.1e-6)),
    'ExtendedKalmanTest': (ExtendedKalmanTest, dict(
        trade_usd=2000.0, k=2.0, a=7.0, warmup=180, reversion=False,
        q_level=0.1e-3, q_vel=0.1e-6, q_acc=0.1e-6,
        k_exit=0.0,       # TUNED: dead zone OFF (it cut reversions short, ~7x worse)
        min_hold=1,       # TUNED: no minimum hold; delaying reversals hurt
        cost_mult=0.0,    # TUNED: cost gate off (no measurable benefit)
        drag_m=0.001,     # TUNED: drag sweet spot (1e-4 over-damps, off is weaker)
        c_d_window=180,
        trend_bias=False)),  # dropped: trend gate fights a mean-reversion entry
    'AccelerationKalmanTest': (AccelerationKalmanTest, dict(
        trade_usd=2000.0, k=2.0, warmup=180, reversion=False,
        q_level=0.2e-3, q_vel=0.2e-6, q_acc=0.2e-9)),
    'EMNonLinTest': (EMNonLinTest, dict(
        trade_usd=2000.0, k=2.5, warmup=180,
        reversion=True,   # TUNED: fade the innovation. The filter's prediction
                          # lags price (innov ACF ~0.97), so a band break marks
                          # overextension, not momentum — 1h sweep: fade
                          # +8.6k..+16.3k across k=1.5..4, follow mirror-negative.
                          # (Pre-2026-07-16 code had the mapping swapped; old
                          # reversion=False runs were ALREADY fading.)
        q_level=0.1e-3, q_vel=0.1e-6, q_acc=0.1e-9, d0=0.0,
        d_max=0.5,        # TUNED: bounds the drag-induced ringing (complex
                          # eigenvalues 1±i·sqrt(2d|v|/z)). 1h sweep: pnl +22.5k
                          # / Sharpe 2.34 vs +17.6k / 1.69 at d_max=5; d=0
                          # removes ringing but collapses the edge (+6k).
        em_window=720, em_interval=1440,
        em_iters=2,       # TUNED: 3rd EM iter overfits drag into instability
        max_hold=0,       # OFF. Trade-replay suggested a 14d time stop helps,
                          # but the REAL backtest disagrees: after a forced
                          # flatten the still-stretched innovation re-enters
                          # immediately, so it just churns (A/B: +16.3k -> -1.7k).
                          # Set ~336 (1h) only if minimizing drawdown matters
                          # more than pnl (maxDD 12.5% -> 6.7%).
        long_only=True,   # TUNED: real A/B +17.6k / Sharpe 1.69 / win 65.5%
                          # vs +16.3k / 1.41 with shorts — shorts added risk
                          # for nothing (all 10 worst trades were shorts).
        dead=300)),
    'EMACrossShortTest': (EMACrossShortTest, dict(
        trade_usd=2000.0, fast_ema_period=60, slow_ema_period=120)),
    'EMNonLinScaled': (EMNonLinScaled, dict(
        trade_usd=2000.0, k=2.5, warmup=180, reversion=True,
        q_level=0.1e-3, q_vel=0.1e-6, q_acc=0.1e-9, d0=0.0, d_max=0.5,
        em_window=720, em_interval=1440, em_iters=2, max_hold=0,
        long_only=False,   # shorts ON so an opposite signal FLIPS long->short
                           # (with long_only=True the flip degrades to flatten)
        dead=300,
        # scale-in on by default for THIS variant (it's the whole point).
        # Budget mode: deploy up to $2k TOTAL across the ramp (first tranche +
        # 4 adds, each 1.5x the last, one every 3% further against the entry).
        # First-tranche size is derived from the budget, so trade_usd is only
        # used if you clear scale_max_usd back to 0.
        scale_in=True, scale_levels=4, scale_step_pct=0.03,
        scale_size_mult=1.5, scale_max_usd=2000.0)),
}

# Every strategy inherits use_trail/trail_pct from PortfolioStrategy.params
# (strategy_base.py) — list them for every entry (and the module-default
# STRATEGY_PARAMS, used when an unknown strategy name falls back to it) so
# they show up as normal editable fields in the dashboard's per-strategy
# "Configure" params list, same as k/warmup/etc., instead of being hidden
# behind a dedicated checkbox.
# Strategies that opt OUT of the trailing stop entirely — the knobs aren't
# added to their Configure params, so a run can't enable use_trail for them.
# EMNonLinScaled averages DOWN as price moves against it; a protective trailing
# stop sits right in that path and would flatten the ramp exactly when it's
# adding, so the two are mutually exclusive by design.
_NO_TRAIL_STRATEGIES = {'EMNonLinScaled'}
STRATEGY_PARAMS.setdefault('use_trail', False)
STRATEGY_PARAMS.setdefault('trail_pct', 0.05)
for _name, (_cls, _params) in STRATEGIES.items():
    if _name in _NO_TRAIL_STRATEGIES:
        continue
    _params.setdefault('use_trail', False)
    _params.setdefault('trail_pct', 0.05)
del _name, _cls, _params


def _resolve_run_config(strategy=None, params=None, days=None):
    """(cls, effective_params, days) for a run. Unknown strategy names fall
    back to the module default; param values are coerced to the default's type
    and unknown keys dropped (passing a foreign param straight into backtrader
    raises 'unexpected keyword argument' deep in strategy __init__)."""
    if strategy and strategy in STRATEGIES:
        cls, eff = STRATEGIES[strategy][0], dict(STRATEGIES[strategy][1])
    else:
        cls, eff = STRATEGY, dict(STRATEGY_PARAMS)
    for k, v in (params or {}).items():
        if k == 'diag':
            # always accepted (drives the additional-data CSVs) even though
            # the tuned-default dicts don't list it
            eff['diag'] = v if isinstance(v, bool) else str(v).lower() in ('1', 'true', 'yes', 'on')
            continue
        if k == 'pyfolio_report':
            # run-time only flag (gates the tear-sheet step in run()); not a
            # strategy kwarg, popped back out before cerebro.addstrategy()
            eff['pyfolio_report'] = v if isinstance(v, bool) else str(v).lower() in ('1', 'true', 'yes', 'on')
            continue
        if k not in eff:
            print('[backtest] ignoring unknown param %r for %s' % (k, cls.__name__))
            continue
        d = eff[k]
        try:
            if isinstance(d, bool):
                v = v if isinstance(v, bool) else str(v).lower() in ('1', 'true', 'yes', 'on')
            elif isinstance(d, int):
                v = int(float(v))
            elif isinstance(d, float):
                v = float(v)
        except (TypeError, ValueError):
            print('[backtest] bad value %r for param %r — keeping default %r' % (v, k, d))
            continue
        eff[k] = v
    if not eff.get('use_trail'):
        # trail_pct is meaningless without use_trail — force it to None so a
        # saved run's params never show a stale/misleading percent when the
        # feature was actually off (strategy_base.py already ignores
        # trail_pct whenever use_trail is False, so this is display-only).
        eff['trail_pct'] = None
    try:
        days = int(days) if days else BACKTEST_DAYS
    except (TypeError, ValueError):
        days = BACKTEST_DAYS
    return cls, eff, max(1, days)

BACKTEST_DAYS = 7
BACKTEST_END = None   # 'YYYY-MM-DD HH:MM' (UTC), or None = latest available bar


def _status(**kw):
    os.makedirs(REPORTS, exist_ok=True)
    kw['ts'] = time.time()
    with open(STATUS_PATH, 'w', encoding='utf-8') as f:
        json.dump(kw, f)


def _refresh_full_csvs(conn, timeframe, need_start=None):

    import db
    data_dir, suffix, _ = _tf_cfg(timeframe)
    os.makedirs(data_dir, exist_ok=True)
    bar_s = 3600 if timeframe == '1h' else 60
    tol = timedelta(seconds=2 * bar_s)
    refreshed = skipped = 0
    for sym, s_min, s_max, _cnt in db.per_symbol_span(conn, timeframe):
        want_start = max(s_min, need_start) if need_start else s_min
        path = os.path.join(data_dir, '%s-%s.csv' % (sym, suffix))
        try:
            first, last = _csv_span(path)
            fresh = (abs((last.replace(tzinfo=timezone.utc) - s_max).total_seconds()) < 2 * bar_s
                     and first.replace(tzinfo=timezone.utc) <= want_start + tol)
        except (OSError, ValueError, IndexError):
            fresh = False
        if fresh:
            skipped += 1
            continue
        db.export_bars_csv(conn, sym, want_start, s_max, path, interval=timeframe)
        refreshed += 1
    print('[backtest] csv cache (%s): %d refreshed, %d already current'
          % (timeframe, refreshed, skipped))


def _prepare_csvs_from_db(days=None, timeframe='1m'):
 
    try:
        import db
        conn = db.get_conn()
    except Exception as e:
        print('[backtest] DB unavailable (%s) — using existing CSVs in datas/' % e)
        return None
    try:
        mn, mx = db.bars_span(conn, timeframe)
        if mn is None:
            raise ValueError('database has no bars — run fetch_binance_csv.py first')
        naive = lambda t: t.astimezone(timezone.utc).replace(tzinfo=None)
        avail_first, avail_last = naive(mn), naive(mx)

        if BACKTEST_END:
            end = datetime.strptime(BACKTEST_END, '%Y-%m-%d %H:%M')
        else:
            end = avail_last
        start = end - timedelta(days=days or BACKTEST_DAYS)

        tol = timedelta(hours=1)
        if start < avail_first - tol or end > avail_last + tol:
            raise ValueError(
                '%s -> %s UTC period not present in data '
                '(available %s -> %s), aborting run'
                % (start, end, avail_first, avail_last))

        if not os.environ.get('JOB_SKIP_REFRESH'):
            _refresh_full_csvs(conn, timeframe,
                               need_start=start.replace(tzinfo=timezone.utc))

        # symbols that only enter mid-window (listed late) — informational
        late = [(sym, s_min) for sym, s_min, _mx2, _c in db.per_symbol_span(conn, timeframe)
                if (s_min - start.replace(tzinfo=timezone.utc)).total_seconds() > 86400]
        if late:
            late.sort(key=lambda x: x[1])
            preview = ', '.join('%s(%s)' % (s, d.strftime('%Y-%m-%d')) for s, d in late[:10])
            print('[backtest] NOTE: %d symbol(s) enter after the window start: %s%s'
                  % (len(late), preview, '' if len(late) <= 10 else ' ...'))
        return start, end
    finally:
        conn.close()


def _csv_span(path):

    with open(path, 'rb') as f:
        f.readline()                              # header
        first = f.readline().split(b',')[0].decode()
        f.seek(0, os.SEEK_END)
        f.seek(max(0, f.tell() - 4096))
        last = f.read().splitlines()[-1].split(b',')[0].decode()
    fmt = '%Y-%m-%d %H:%M:%S'
    return datetime.strptime(first, fmt), datetime.strptime(last, fmt)


def _resolve_window(days=None, timeframe='1m'):

    spans = []
    _dd, _suf, _ = _tf_cfg(timeframe)
    for p in glob.glob(os.path.join(_dd, '*-%s.csv' % _suf)):
        try:
            spans.append(_csv_span(p))
        except Exception:
            pass                                   # empty/corrupt file
    if not spans:
        raise ValueError('no usable CSVs in datas/ — run fetch_binance_csv.py first')
    avail_first = min(s[0] for s in spans)
    avail_last = max(s[1] for s in spans)

    if BACKTEST_END:
        end = datetime.strptime(BACKTEST_END, '%Y-%m-%d %H:%M')
    else:
        end = avail_last
    start = end - timedelta(days=days or BACKTEST_DAYS)

    tol = timedelta(hours=1)   # first/last bars may sit just inside the edge
    if start < avail_first - tol or end > avail_last + tol:
        raise ValueError(
            '%s -> %s UTC period not present in data '
            '(available %s -> %s), aborting run'
            % (start, end, avail_first, avail_last))
    return start, end


def _load_feeds(cerebro, fromdate, todate, timeframe='1m'):
    symbols = []
    data_dir, suffix, comp = _tf_cfg(timeframe)
    for path in sorted(glob.glob(os.path.join(data_dir, '*-%s.csv' % suffix))):
        sym = os.path.basename(path)[:-len('-%s.csv' % suffix)]
        data = bt.feeds.GenericCSVData(
            dataname=path,
            dtformat='%Y-%m-%d %H:%M:%S',
            timeframe=bt.TimeFrame.Minutes,
            compression=comp,
            datetime=0, open=1, high=2, low=3, close=4, volume=5,
            openinterest=-1,
            name=sym,
            fromdate=fromdate,
            todate=todate,
        )
        cerebro.adddata(data)
        symbols.append(sym)
    return symbols


def _epoch(dt):
  
    return int(dt.replace(tzinfo=timezone.utc).timestamp())

def _resample_returns(rets, bars_per_period=60*24):

    out = []
    acc = 1.0
    count = 0
    for r in rets:
        if r != r:          # skip NaN
            continue
        acc *= (1.0 + r)
        count += 1
        if count == bars_per_period:
            out.append(acc - 1.0)
            acc = 1.0
            count = 0
    return out

def _risk_metrics(rets, bars_per_day=1440):

    import pandas as pd

    r_minute = [x for x in rets if x == x]        # drop NaN
    r = _resample_returns(r_minute, bars_per_day)
    n = len(r)
    if n < 2:
        return {}
    T = MINUTES_PER_YEAR/60 *1/24

    mu = sum(r) / n
    var = sum((x - mu) ** 2 for x in r) / (n - 1)      # ddof=1, matches empyrical
    sd = math.sqrt(var)

    cum = 1.0
    for x in r_minute:
        cum *= (1.0 + x)
    cum -= 1.0

    arith = mu * T
    try:
        cagr = (1.0 + cum) ** (T / n) - 1.0 if cum > -1.0 else float('nan')
    except OverflowError:
        cagr = float('inf')

    #log sharpe calculation
    import numpy as np
    log_returns = np.log(1+np.array(r))
    log_risk_free_rate = np.log(RISK_FREE_ANNUAL + 1)

    log_returns_annualised = np.mean(log_returns) * T
    ann_vol_log = np.std(log_returns, ddof = 1) * np.sqrt(T)

    ann_vol_arithmetic = sd * math.sqrt(T)

    out = {'cumulative_return': cum,
           'annual_return_cagr': cagr,
           'annual_volatility_returns': ann_vol_arithmetic,
           'annual_volatility_log_returns':ann_vol_log,
           'sharpe_arithmetic': None,
           'sharpe_log_returns': None}
    if ann_vol_arithmetic > 0:
        out['sharpe_arithmetic'] = (arith - RISK_FREE_ANNUAL) / ann_vol_arithmetic
        out['sharpe_log_returns'] = (log_returns_annualised - log_risk_free_rate) / ann_vol_log
    return out


_ADD_PALETTE = ['#58a6ff', '#e8834c', '#5cc98a', '#b58cf0', '#e5566a',
                '#4fc6c0', '#d29922', '#8b96a5', '#ff7eb6']


def _additional_lines(rows):

    if not rows:
        return []
    keys = list(rows[0].keys())
    groups = [
        ('innovation', [k for k in ('innov', 'band') if k in keys]),
        ('state', [k for k in keys if k in ('x1', 'x2')]),
        ('covariance', [k for k in keys if k.startswith('P')]),
    ]
    out, ci = [], 0
    for gname, ks in groups:
        for k in ks:
            pts = [{'time': r['t'], 'value': round(float(r[k]), 10)}
                   for r in rows
                   if r.get(k) is not None and r[k] == r[k]]
            if len(pts) < 2:
                continue
            out.append({'name': k, 'color': _ADD_PALETTE[ci % len(_ADD_PALETTE)],
                        'kind': 'additional', 'group': gname, 'points': pts})
            ci += 1
    return out


def _dump_chartdata(strat):

    os.makedirs(CHARTDATA_DIR, exist_ok=True)
    saved = []
    total = len(strat.datas)

    try:
        lines_by_symbol = strat.build_chart_lines() or {}
    except Exception as e:
        print('[backtest] build_chart_lines failed (%s) — charts get no overlays' % e)
        lines_by_symbol = {}

    # Positions grouped per symbol (oldest first) for the per-stock table.
    pos_by_symbol = {}
    for t in strat.trade_log:
        pos_by_symbol.setdefault(t['symbol'], []).append(t)
    for lst in pos_by_symbol.values():
        lst.sort(key=lambda t: t.get('entry_dt') or '')

    for i, d in enumerate(strat.datas):
        times = [_epoch(bt.num2date(v)) for v in d.datetime.array]
        o, h = list(d.open.array), list(d.high.array)
        low, c = list(d.low.array), list(d.close.array)

        candles = [{'time': t, 'open': o[j], 'high': h[j],
                    'low': low[j], 'close': c[j]}
                   for j, t in enumerate(times)]

        markers = []
        for e in strat.executed[d._name]:
            buy = e['side'] == 'BUY'
            markers.append({
                'time': _epoch(datetime.fromisoformat(e['dt'])),
                'position': 'belowBar' if buy else 'aboveBar',
                'color': '#26a69a' if buy else '#ef5350',
                'shape': 'arrowUp' if buy else 'arrowDown',
                'text': e['side'],
            })
        markers.sort(key=lambda m: m['time'])

        lines = list(lines_by_symbol.get(d._name, []))
        for L in lines:
            L.setdefault('kind', 'main')
        lines += _additional_lines(strat._diag.get(d._name, []))
        cd = {'symbol': d._name,
              'candles': candles,
              'lines': lines,
              'markers': markers,
              'positions': pos_by_symbol.get(d._name, [])}
        with open(os.path.join(CHARTDATA_DIR, '%s.json' % d._name), 'w',
                  encoding='utf-8') as f:
            json.dump(cd, f)
        saved.append(d._name)

        if (i + 1) % 5 == 0 or (i + 1) == total:
            _status(state='running', phase='charts', done=i + 1, total=total)
    return saved


_POS_COLS = ['side', 'size', 'entry_signal_dt', 'entry_dt',
             'exit_signal_dt', 'exit_dt', 'entry_price',
             'exit_price', 'bars_held', 'pnl', 'pnlcomm',
             'exit_reason', 'trail_stop_price']


def _dump_position_logs(strat, out_dir, per_symbol_dir=None):

    import csv as _csv
    by_symbol = {}
    for t in strat.trade_log:
        by_symbol.setdefault(t['symbol'], []).append(t)
    for lst in by_symbol.values():
        lst.sort(key=lambda t: t.get('entry_dt') or '')

    if per_symbol_dir:
        os.makedirs(per_symbol_dir, exist_ok=True)
        for sym, rows in by_symbol.items():
            with open(os.path.join(per_symbol_dir, '%s.csv' % sym), 'w',
                      newline='', encoding='utf-8') as f:
                w = _csv.DictWriter(f, fieldnames=_POS_COLS, extrasaction='ignore')
                w.writeheader()
                w.writerows(rows)

    # combined file (symbol-prefixed), grouped by symbol then entry time
    combined = sorted(strat.trade_log,
                      key=lambda t: (t['symbol'], t.get('entry_dt') or ''))
    with open(os.path.join(out_dir, 'position_log.csv'), 'w',
              newline='', encoding='utf-8') as f:
        w = _csv.DictWriter(f, fieldnames=['symbol'] + _POS_COLS,
                            extrasaction='ignore')
        w.writeheader()
        w.writerows(combined)
    return len(strat.trade_log)


def _dump_run_config(out_dir, strat, bt_start, bt_end, n_symbols, days=None,
                     timeframe='1m'):
    import csv as _csv
    rows = [('strategy', type(strat).__name__), ('timeframe', timeframe)]
    eff = {k: getattr(strat.params, k) for k in strat.params._getkeys()}
    for k, v in eff.items():
        rows.append(('param.%s' % k, v))
    rows += [
        ('broker.starting_cash', STARTING_CASH),
        ('broker.leverage', LEVERAGE),
        ('broker.commission', COMMISSION),
        ('broker.commission_pct_per_side', '%.4f%%' % (COMMISSION * 100)),
        ('risk_free_annual', RISK_FREE_ANNUAL),
        ('window_start', bt_start.isoformat()),
        ('window_end', bt_end.isoformat()),
        ('backtest_days', days or BACKTEST_DAYS),
        ('symbols', n_symbols),
    ]
    """with open(os.path.join(out_dir, 'config.csv'), 'w',
              newline='', encoding='utf-8') as f:
        w = _csv.writer(f)
        w.writerow(['setting', 'value'])
        w.writerows(rows)"""
    return rows


def _write_report_xlsx(out_dir, config_rows, stats, run_tag):

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception as e:
        print('[backtest] xlsx skipped (%s)' % e)
        return
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = str(run_tag)[:31]
        thin = Side(style='thin', color='000000')
        box = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center')
        stripe = PatternFill('solid', fgColor='D9D9D9')
        red = PatternFill('solid', fgColor='DA9694')
        green = PatternFill('solid', fgColor='A9D08E')
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 26

        def header(row, left, right):
            for col, text in ((1, left), (2, right)):
                c = ws.cell(row=row, column=col, value=text)
                c.font = Font(bold=True)
                c.alignment = center
                c.border = box

        header(1, 'setting', 'value')
        r = 2
        for i, (k, v) in enumerate(config_rows):
            ws.cell(row=r, column=1, value=k)
            c = ws.cell(row=r, column=2, value=v)
            if isinstance(v, bool):
                c.value = 'TRUE' if v else 'FALSE'
                c.alignment = center
            if i % 2 == 0:                       # striped config rows
                ws.cell(row=r, column=1).fill = stripe
                c.fill = stripe
            r += 1

        header(r, 'metric', 'value')
        r += 1
        for k, v in (stats or {}).items():
            name = ws.cell(row=r, column=1, value=k)
            name.font = Font(bold=True)
            name.alignment = center
            name.border = box
            c = ws.cell(row=r, column=2, value=v)
            c.border = box
            if isinstance(v, (int, float)) and v == v:
                c.fill = red if v < 0 else green
            r += 1

        wb.save(os.path.join(out_dir, 'report.xlsx'))
    except Exception as e:
        print('[backtest] per-run xlsx failed: %s' % e)

def _pyfolio_report(strat, out_dir, timeframe='1m', want_images=True):

    os.makedirs(out_dir, exist_ok=True)
    images, stats, error = [], {}, None

    pyf = strat.analyzers.getbyname('pyfolio')
    returns, positions, transactions, gross_lev = pyf.get_pf_items()

    # Always dump the raw items so they can be analyzed elsewhere.
    returns.to_csv(os.path.join(out_dir, 'returns_pyfolio.csv'))
    try:
        transactions.to_csv(os.path.join(out_dir, 'transactions.csv'))
    except Exception:
        pass

    try:
        import pandas as pd
        import pyfolio as pf

        try:
            import empyrical as ep
            # returns is pyfolio's own daily-resampled series (see the
            # analyzer's default timeframe/compression), so annualization is
            # always 365 observations/year regardless of the backtest's bar size.
            MPY = 365
            rf = RISK_FREE_ANNUAL / MPY  # per-observation rate
            r = returns.dropna()

            rm = _risk_metrics(list(r.values), 1)
            raw = {
                'Cumulative return': rm.get('cumulative_return'),
                'Annualised return (CAGR)': rm.get('annual_return_cagr'),
                'Annual volatility (Return)': rm.get('annual_volatility_returns'),
                'Annual volatility (Log Return)': rm.get('annual_volatility_log_returns'),
                'Sharpe ratio (Returns)': rm.get('sharpe_arithmetic'),
                'Sharpe ratio (log returns)': rm.get('sharpe_log_returns'),

                'Sortino ratio':     ep.sortino_ratio(r, annualization=MPY,
                                                      required_return=rf),
                'Calmar ratio':      ep.calmar_ratio(r, annualization=MPY),
                'Max drawdown':      ep.max_drawdown(r),
                'Stability':         ep.stability_of_timeseries(r),
                'Tail ratio':        ep.tail_ratio(r),
                'Daily VaR (5%)':    ep.value_at_risk(r),
                'Skew':              float(r.skew()),
                'Kurtosis':          float(r.kurtosis()),
            }
            stats = {k: (None if v is None or (isinstance(v, float) and v != v)
                         else round(float(v), 6))
                     for k, v in raw.items()}

            # Trade counters (integers, kept out of the float rounding above).
            wins = sum(1 for t in strat.trade_log if t['pnlcomm'] > 0)
            n_trades = len(strat.trade_log)
            stats['Total trades'] = n_trades
            stats['Winning trades'] = wins
            stats['Losing trades'] = n_trades - wins
            stats['Win rate %'] = (round(100.0 * wins / n_trades, 2)
                                   if n_trades else None)

            """(pd.Series(stats, dtype=object).rename('value')
             .to_csv(os.path.join(out_dir, 'perf_stats.csv'),
                     index_label='metric'))"""
        except Exception as e:
            error = 'stats: %s' % e

        if want_images:
            try:
                plt.close('all')
                fig = pf.create_returns_tear_sheet(returns, return_fig=True)
                fig.savefig(os.path.join(out_dir, 'returns_tear_sheet.png'),
                            dpi=90, bbox_inches='tight')
                images.append('returns_tear_sheet.png')
                plt.close('all')
            except Exception as e:
                error = (error + ' | ' if error else '') + 'returns_tear_sheet: %s' % e
    except ImportError:
        error = 'pyfolio not installed (pip install pyfolio-reloaded)'

    return images, stats, error


def _dget(node, *keys, default=0):
    cur = node
    for k in keys:
        try:
            cur = cur[k]
        except (KeyError, TypeError, IndexError):
            return default
    return cur if cur or cur == 0 else default


def run(strategy=None, params=None, days=None, timeframe='1m', name=None):
    started = time.time()
    # optional friendly label — doesn't need to be unique; defaults to the
    # run tag (the archive folder name, generated below once known) when blank
    name = str(name).strip() if name else None
    cls, eff_params, days = _resolve_run_config(strategy, params, days)
    # run-time only; not a strategy constructor kwarg
    want_pyfolio_report = eff_params.pop('pyfolio_report', False)
    timeframe = timeframe if timeframe in ('1m', '1h') else '1m'
    # Per-bar filter internals ("additional data" CSVs) are recorded when the
    # diag param is on — the dashboard's builder exposes it as a checkbox
    # (defaults on for 1h; heavy on full-universe 1m runs: GBs of RAM).

    try:
        _status(state='running', phase='loading',
                strategy=cls.__name__, days=days, timeframe=timeframe)
        os.makedirs(REPORTS, exist_ok=True)
        shutil.rmtree(CHARTDATA_DIR, ignore_errors=True)
        for stale in ('trades.csv', 'positions.csv'):
            try:
                os.remove(os.path.join(REPORTS, stale))
            except OSError:
                pass

        try:
            window = _prepare_csvs_from_db(days, timeframe)
            if window is None:
                window = _resolve_window(days, timeframe)
            bt_start, bt_end = window
        except ValueError as e:
            print('[backtest] %s' % e)
            _status(state='error', error=str(e))
            return
        print('[backtest] window: %s -> %s UTC' % (bt_start, bt_end))

        cerebro = bt.Cerebro()  # stdstats=True -> BuySell arrows on the plots
        cerebro.broker.setcash(STARTING_CASH)
        cerebro.broker.setcommission(commission=COMMISSION, leverage=float(LEVERAGE))

        symbols = _load_feeds(cerebro, bt_start, bt_end, timeframe)
        if not symbols:
            _status(state='error',
                    error='no CSVs in datas/ — run fetch_binance_csv.py first')
            return
        print('[backtest] %d feeds loaded' % len(symbols))

#--------------------------------------------------------------------------------------
        cerebro.addstrategy(cls, **eff_params)
#---------------------------------------------------------------------------------------

        cerebro.addanalyzer(bt.analyzers.PyFolio, _name='pyfolio')
        cerebro.addanalyzer(bt.analyzers.TradeAnalyzer, _name='trades')
        cerebro.addanalyzer(bt.analyzers.DrawDown, _name='dd')

        _status(state='running', phase='engine', total=len(symbols))

        strat = cerebro.run()[0]
        end_value = cerebro.broker.getvalue()
        print('[backtest] engine done, end value %.2f' % end_value)

        ta = strat.analyzers.trades.get_analysis()
        dd = strat.analyzers.dd.get_analysis()
        # "Positions" = closed position lifecycles (flat -> open -> flat). A
        # scale-in ramp with several add-on tranches is still ONE position.
        n_closed = _dget(ta, 'total', 'closed')
        n_won = _dget(ta, 'won', 'total')
        n_lost = _dget(ta, 'lost', 'total')
        # "Trades" = individual order fills (every entry, add-on, and exit —
        # what actually executed on the exchange). >= n_closed always; equal
        # to it only when nothing ever adds to an open position.
        n_fills = sum(len(v) for v in strat.executed.values())

        # Unrealised (mark-to-market) PnL of positions STILL OPEN at the end
        # of the backtest — realised pnlcomm in the per-symbol logs only
        # covers CLOSED positions, so on a run that ends holding open
        # positions the per-symbol realised sum won't add up to end_value.
        # unrealised = (last_close - avg_entry_price) * signed_size, gross of
        # the not-yet-paid exit commission (same basis backtrader uses to
        # mark open positions into getvalue()).
        unrealised_by_symbol = {}
        for d in strat.datas:
            pos = strat.getposition(d)
            if pos.size and len(d):
                last_close = d.close[0]
                unrealised_by_symbol[d._name] = round(
                    (last_close - pos.price) * pos.size, 6)
        unrealised_total = round(sum(unrealised_by_symbol.values()), 4)

        pyf = strat.analyzers.getbyname('pyfolio')
        pf_returns, _pf_pos, _pf_txn, _pf_lev = pyf.get_pf_items()
        rets = list(pf_returns.dropna().values)

        rm = _risk_metrics(rets, 1)
        rnd = lambda k, d=4: (None if rm.get(k) is None or rm.get(k) != rm.get(k)
                              else round(rm[k], d))

        summary = {
            'start_cash': STARTING_CASH,
            'end_value': round(end_value, 2),
            'pnl': round(end_value - STARTING_CASH, 2),
            'return_pct': round((end_value / STARTING_CASH - 1) * 100, 4),
            'trades_closed': n_closed,     # "Positions" in the UI — closed position lifecycles
            'trades_total': n_fills,       # "Trades" in the UI — individual order fills
            'unrealised_pnl': unrealised_total,   # MTM of positions open at end
            'open_positions': len(unrealised_by_symbol),
            'won': n_won,
            'lost': n_lost,
            'win_rate_pct': round(100.0 * n_won / n_closed, 2) if n_closed else None,
            'max_drawdown_pct': round(_dget(dd, 'max', 'drawdown', default=0.0), 4),
            'annual_return_cagr': rnd('annual_return_cagr'),
            'annual_volatility_returns': rnd('annual_volatility_returns'),
            'annual_volatility_log_returns': rnd('annual_volatility_log_returns'),
            'sharpe_arithmetic': rnd('sharpe_arithmetic'),
            'sharpe_log_returns': rnd('sharpe_log_returns'),
            'symbols': len(symbols),
        }

        # ---- per-symbol -------------------------------------------------
        per_symbol = {}
        for t in strat.trade_log:
            s = per_symbol.setdefault(t['symbol'],
                                      {'trades': 0, 'pnl': 0.0, 'won': 0})
            s['trades'] += 1
            s['pnl'] += t['pnlcomm']
            s['won'] += 1 if t['pnlcomm'] > 0 else 0
        for sym in symbols:
            per_symbol.setdefault(sym, {'trades': 0, 'pnl': 0.0, 'won': 0})
        for sym, s in per_symbol.items():
            s['pnl'] = round(s['pnl'], 4)
            s['unrealised'] = unrealised_by_symbol.get(sym, 0.0)

        # sidecar the raw open-position MTM so the per-symbol/class
        # contribution endpoints can surface an "Unrealised" column alongside
        # the realised numbers they read from the position CSVs
        try:
            with open(os.path.join(REPORTS, 'unrealised.json'), 'w',
                      encoding='utf-8') as f:
                json.dump(unrealised_by_symbol, f)
        except OSError as e:
            print('[backtest] unrealised.json write failed: %s' % e)

        # ---- chart data (candles + EMAs + fills for interactive charts) -
        _status(state='running', phase='charts', done=0, total=len(symbols))
        charts = _dump_chartdata(strat)
        # symbols whose chartdata carries additional (diag) series
        extras = sorted(set(charts) & {s for s, rows in strat._diag.items() if rows})
        print('[backtest] chart data written for %d symbols' % len(charts))

        # ---- per-run report archive (pyfolio + position logs) -----------
        _status(state='running', phase='pyfolio')
        # pid suffix keeps archive tags unique when parallel jobs finish in
        # the same second
        run_tag = time.strftime('%Y%m%d-%H%M%S') + '-p%d' % os.getpid()
        effective_name = name or run_tag
        run_dir = os.path.join(TEST_DATA_DIR, run_tag)
        pf_images, pf_stats, pf_error = _pyfolio_report(
            strat, run_dir, timeframe, want_images=want_pyfolio_report)
        import pandas as pd
        # accumulated log -> archive; per-symbol files -> the job/run folder
        n_pos = _dump_position_logs(strat, run_dir,
                                    per_symbol_dir=os.path.join(REPORTS, 'positions'))
        config_rows = _dump_run_config(run_dir, strat, bt_start, bt_end,
                                       len(symbols), days, timeframe)
        _write_report_xlsx(run_dir, config_rows, pf_stats, run_tag)   # config + perf_stats
        print('[backtest] report done (%s) -> test_data/%s (%d positions)'
              % (pf_error or 'ok', run_tag, n_pos))

        equity = strat.equity
        if len(equity) > 3000:  # thin for the dashboard
            step = len(equity) // 3000 + 1
            equity = equity[::step] + [strat.equity[-1]]

        # Full position log to CSV (one row per closed position, grouped by
        # symbol then entry time) + capped list in the JSON.
        import csv as _csv
        cols = ['symbol'] + _POS_COLS
        positions = sorted(strat.trade_log,
                           key=lambda t: (t['symbol'], t.get('entry_dt') or ''))
        with open(os.path.join(REPORTS, 'positions.csv'), 'w',
                  newline='', encoding='utf-8') as f:
            w = _csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
            w.writeheader()
            w.writerows(positions)

        results = {
            'generated': datetime.now(timezone.utc).isoformat(),
            'name': effective_name,
            'run_tag': run_tag,
            'params': dict(eff_params,
                           strategy=cls.__name__,
                           timeframe=timeframe,
                           leverage=LEVERAGE,
                           window_start=bt_start.isoformat(),
                           window_end=bt_end.isoformat()),
            'summary': summary,
            'per_symbol': per_symbol,
            'charts': charts,
            'extras': extras,     # symbols with an additional-data CSV
            'equity': equity,
            'trades': strat.trade_log[-2000:],
            'pyfolio': {'images': pf_images, 'stats': pf_stats,
                        'error': pf_error,
                        'dir': 'reports/test_data/%s' % run_tag},
        }

        _status(state='running', phase='database')
        try:
            import db
            conn = db.get_conn()
            db.init_schema(conn)
            run_id = db.save_run(conn, results,
                                 strat.trade_log, strat.equity)
            conn.close()
            results['db'] = {'saved': True, 'run_id': run_id}
            print('[backtest] saved to postgres as run %d' % run_id)
        except Exception as e:
            results['db'] = {'saved': False, 'error': str(e)}
            print('[backtest] postgres save failed (file outputs intact): %s' % e)

        with open(RESULTS_PATH, 'w', encoding='utf-8') as f:
            json.dump(results, f)


        snapshot = {k: results[k] for k in
                    ('generated', 'name', 'params', 'summary', 'per_symbol', 'equity', 'pyfolio')}
        snapshot['run_tag'] = run_tag
        try:
            with open(os.path.join(run_dir, 'run.json'), 'w', encoding='utf-8') as f:
                json.dump(snapshot, f)
        except OSError as e:
            print('[backtest] run.json snapshot failed: %s' % e)

        _status(state='done', elapsed=round(time.time() - started, 1),
                pnl=summary['pnl'], trades=n_closed)
        print('[backtest] done in %.1fs, pnl %.2f, %d trades'
              % (time.time() - started, summary['pnl'], n_closed))
    except Exception as e:
        print('[backtest] FAILED: %s\n%s' % (e, traceback.format_exc()))
        _status(state='error', error=str(e))


if __name__ == '__main__':
    run()
