
import asyncio
import importlib
import json
from collections import deque
from datetime import datetime, timezone
import os
import shutil
import sys
import threading

import uvicorn
from fastapi import FastAPI, Request
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles

import _paths                      # noqa: F401  (sys.path: ROOT/backend/strategies)
import db
import strategy_base
import EMACrossShortTest
import KalmanFilter
import ExtendedKalmanFilter
import AccelerationKalmanFilter
import EMAlgoTest
import EMAlgoNonLinTest
import run_backtest
import symbol_classes
import pnl_stats
import LSTM

BASE = _paths.ROOT                 # data/artifacts live at project root
REPORTS = os.path.join(BASE, 'reports')
os.makedirs(REPORTS, exist_ok=True)

# ---- server log ring buffer -------------------------------------------------
# Everything printed by the server and its worker threads (backtests, sync,
# report builds) is teed into this buffer and served at /api/logs so the
# dashboard's log window can show it (warnings about excluded symbols, run
# tracebacks, etc.) without shell access.
_LOG_BUF: deque = deque(maxlen=4000)     # (id, 'HH:MM:SS', line)
_log_lock = threading.Lock()
_log_seq = 0
# access-log lines from the dashboard's own polling would flood the window
# (the log poll would log itself every 2s) — drop them at capture time
_LOG_SKIP = ('/api/logs', '/api/status', '/api/sync/status',
             '/api/report/status', '/api/results')


class _LogTee:
    """File-like wrapper: pass writes through and collect complete lines."""
    def __init__(self, orig):
        self._orig = orig
        self._part = ''

    def write(self, s):
        try:
            self._orig.write(s)
        except Exception:
            pass
        global _log_seq
        with _log_lock:
            self._part += s
            while '\n' in self._part:
                line, self._part = self._part.split('\n', 1)
                if line.strip() and not any(p in line for p in _LOG_SKIP):
                    _log_seq += 1
                    _LOG_BUF.append((_log_seq,
                                     datetime.now().strftime('%H:%M:%S'),
                                     line.rstrip()))

    def flush(self):
        try:
            self._orig.flush()
        except Exception:
            pass

    def isatty(self):
        return False


sys.stdout = _LogTee(sys.stdout)
sys.stderr = _LogTee(sys.stderr)

app = FastAPI(title='Backtrader Dashboard')
app.mount('/reports', StaticFiles(directory=REPORTS), name='reports')
# Vendored JS (lightweight-charts) so the dashboard needs no internet/CDN.
app.mount('/vendor', StaticFiles(directory=os.path.join(BASE, 'vendor')),
          name='vendor')
# KF/EKF/EM comparison report (build_report.py output) on the same server.
REPORT_OUT = os.path.join(BASE, 'report_out')
os.makedirs(REPORT_OUT, exist_ok=True)
app.mount('/report', StaticFiles(directory=REPORT_OUT, html=True), name='report')

_run_thread: threading.Thread | None = None
_report_thread: threading.Thread | None = None
_report_state: dict = {'state': 'idle'}
_sync_thread: threading.Thread | None = None
_sync_state: dict = {'state': 'idle'}
_console_proc = None                    # persistent console_worker.py subprocess
_console_lock = threading.Lock()


def _read_json(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


# ---- per-run favourites + notes -----------------------------------------
# A tiny sidecar store keyed by a stable run id (job id, archive tag, or a
# history row's DB id — whatever the client passes). Survives restarts; never
# deleted with a job folder, so a note outlives the artifacts.
_RUNMETA_PATH = os.path.join(REPORTS, 'run_meta.json')
_runmeta_lock = threading.Lock()

def _runmeta_load():
    return _read_json(_RUNMETA_PATH) or {}

def _runmeta_save(data):
    os.makedirs(REPORTS, exist_ok=True)
    tmp = _RUNMETA_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f)
    os.replace(tmp, _RUNMETA_PATH)


@app.get('/api/runmeta')
def get_runmeta():
    """Whole favourites/notes map: {id: {favorite: bool, note: str}}."""
    return _runmeta_load()


@app.post('/api/runmeta')
async def set_runmeta(request: Request):
    """Merge one run's meta. Body: {id, favorite?, note?}. Omitted fields are
    left unchanged; note='' clears the note; an entry with neither a favourite
    nor a note is dropped to keep the file tidy."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    rid = str(payload.get('id') or '').strip()
    if not rid:
        return JSONResponse({'error': 'missing id'}, status_code=400)
    with _runmeta_lock:
        data = _runmeta_load()
        cur = dict(data.get(rid) or {})
        if 'favorite' in payload:
            cur['favorite'] = bool(payload['favorite'])
        if 'note' in payload:
            note = str(payload['note'] or '').strip()
            if note:
                cur['note'] = note
            else:
                cur.pop('note', None)
        if cur.get('favorite') or cur.get('note'):
            data[rid] = cur
        else:
            data.pop(rid, None)
        _runmeta_save(data)
    return {'ok': True, 'meta': data.get(rid) or {}}


@app.get('/')
def index():
    return FileResponse(os.path.join(BASE, 'dashboard.html'))


@app.get('/api/results')
def results():
    # Prefer the most recently finished (non-dismissed) job; fall back to the
    # legacy single-run results file.
    with _jobs_lock:
        done = [m for m in _jobs.values()
                if m['state'] == 'done' and not m.get('dismissed')]
    if done:
        jid = max(done, key=lambda m: m.get('started') or m['created'])['id']
        res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
        if res is not None:
            res['chartbase'] = '/reports/jobs/%s/chartdata' % jid
            res['job_id'] = jid
            return res
    data = _read_json(run_backtest.RESULTS_PATH)
    if data is None:
        return JSONResponse({'error': 'no results yet — run a backtest'},
                            status_code=404)
    return data


@app.get('/api/status')
def status():
    running = _run_thread is not None and _run_thread.is_alive()
    data = _read_json(run_backtest.STATUS_PATH) or {'state': 'idle'}
    # If the runner crashed hard before writing an error status, don't report
    # a stale 'running' state forever.
    if data.get('state') == 'running' and not running:
        data['state'] = 'error'
        data.setdefault('error', 'runner stopped unexpectedly')
    data['thread_alive'] = running
    return data


@app.get('/api/strategies')
def strategies():
    """Strategy picker data: every registered strategy with its tuned default
    params, plus the module default selection and backtest days. Reloads
    from disk first (like /api/run) so param/registry edits show up in the
    builder immediately, without restarting the server."""
    _reload_modules()
    return {
        'default': run_backtest.STRATEGY.__name__,
        'days': run_backtest.BACKTEST_DAYS,
        'timeframe': '1m',
        'timeframes': ['1m', '1h'],
        'strategies': {name: params
                       for name, (_cls, params) in run_backtest.STRATEGIES.items()},
    }


@app.post('/api/run')
async def run(request: Request):
    global run_backtest
    if _run_thread is not None and _run_thread.is_alive():
        return JSONResponse({'error': 'backtest already running'},
                            status_code=409)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    payload = payload if isinstance(payload, dict) else {}
    err = _reload_modules()
    if err:
        return JSONResponse({'error': err}, status_code=500)
    return _start_run(payload)


def _reload_modules():
    """Reload every project module from disk so code/param edits (including the
    STRATEGIES registry defaults) take effect without restarting the server.
    Order is dependency-first: base classes and leaf strategy modules before
    run_backtest (which does `from <module> import <Strategy>`); reloading it
    first would capture a stale class, then reloading the strategy module in
    place rebinds that class's globals — mixing old methods with new functions.
    Returns None on success or an error string."""
    global db, strategy_base, EMACrossShortTest, KalmanFilter, EMAlgoTest, EMAlgoNonLinTest, run_backtest, LSTM
    global ExtendedKalmanFilter, AccelerationKalmanFilter
    try:
        db = importlib.reload(db)
        strategy_base = importlib.reload(strategy_base)
        EMACrossShortTest = importlib.reload(EMACrossShortTest)
        KalmanFilter = importlib.reload(KalmanFilter)
        EMAlgoTest = importlib.reload(EMAlgoTest)
        EMAlgoNonLinTest = importlib.reload(EMAlgoNonLinTest)
        ExtendedKalmanFilter = importlib.reload(ExtendedKalmanFilter)
        AccelerationKalmanFilter = importlib.reload(AccelerationKalmanFilter)
        LSTM = importlib.reload(LSTM)
        run_backtest = importlib.reload(run_backtest)
    except Exception as e:
        return 'reload failed: %s' % e
    return None


@app.post('/api/reload')
def reload_modules():
    """Reload strategy modules + run_backtest so edits to code or the STRATEGIES
    registry (default params like EMNonLinTest's `dead`) appear in the picker
    without having to run a backtest first."""
    if _run_thread is not None and _run_thread.is_alive():
        return JSONResponse({'error': 'cannot reload while a backtest is running'},
                            status_code=409)
    err = _reload_modules()
    if err:
        return JSONResponse({'error': err}, status_code=500)
    return {'reloaded': True, 'strategies': list(run_backtest.STRATEGIES.keys())}


def _start_run(payload=None):
    global _run_thread
    payload = payload or {}
    kwargs = {'strategy': payload.get('strategy'),
              'params': payload.get('params'),
              'days': payload.get('days'),
              'timeframe': payload.get('timeframe')}
    _run_thread = threading.Thread(target=run_backtest.run, kwargs=kwargs,
                                   daemon=True)
    _run_thread.start()
    return {'started': True}


@app.post('/api/report')
async def build_comparison_report(request: Request):
    """Rebuild the KF/EKF/EM plotly comparison (build_report.py) in a thread;
    the result is served from /report/. Poll /api/report/status."""
    global _report_thread
    if _report_thread is not None and _report_thread.is_alive():
        return JSONResponse({'error': 'report already building'}, status_code=409)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    payload = payload if isinstance(payload, dict) else {}
    symbol = payload.get('symbol') or None
    # [{'name': <registry name>, 'params': {...}}, ...]; None -> report default
    selections = payload.get('strategies') or None
    if selections is not None and not isinstance(selections, list):
        selections = None
    # 'compare old tests' mode: list of archived run tags to compare instead.
    # dedupe (order-preserving) — the History table can hand back the same
    # tag twice (e.g. a job and its archive both selected, or a stale
    # postgres row still pointing at an already-picked job/archive), and a
    # duplicate tag renders as a duplicated run in the comparison report.
    tags = payload.get('tags') or None
    if tags is not None and not isinstance(tags, list):
        tags = None
    if tags:
        seen = set()
        tags = [t for t in tags if not (t in seen or seen.add(t))]
    timeframe = payload.get('timeframe') if payload.get('timeframe') in ('1m', '1h') else '1m'
    try:
        days = int(payload.get('days') or 0) or None   # None -> full history
    except (TypeError, ValueError):
        days = None

    def _work():
        import time as _time
        _report_state.update(state='building', started=_time.time(), error=None)
        try:
            import build_report
            build_report = importlib.reload(build_report)
            if tags:
                build_report.build_from_folders(tags)
            else:
                # symbol=None/'' -> portfolio report over ALL symbols
                build_report.build(symbol, selections=selections,
                                   timeframe=timeframe, days=days)
            _report_state.update(state='done',
                                 elapsed=round(_time.time() - _report_state['started'], 1))
        except Exception as e:
            _report_state.update(state='error', error=str(e))

    _report_thread = threading.Thread(target=_work, daemon=True)
    _report_thread.start()
    return {'started': True}


@app.get('/api/report/status')
def report_status():
    st = dict(_report_state)
    st['thread_alive'] = _report_thread is not None and _report_thread.is_alive()
    if st.get('state') == 'building' and not st['thread_alive']:
        st['state'] = 'error'
        st.setdefault('error', 'report thread stopped unexpectedly')
    st['ready'] = os.path.exists(os.path.join(REPORT_OUT, 'index.html'))
    return st


def _console_localonly(request):
    """The console executes arbitrary Python — never serve it off-box. If the
    server is ever bound to 0.0.0.0 (e.g. to share the report), this endpoint
    stays localhost-only."""
    host = getattr(request.client, 'host', '')
    return host in ('127.0.0.1', '::1', 'localhost')


def _console_start():
    global _console_proc
    import subprocess
    _console_proc = subprocess.Popen(
        [sys.executable, os.path.join(_paths.BACKEND, 'console_worker.py')],
        stdin=subprocess.PIPE, stdout=subprocess.PIPE,
        text=True, encoding='utf-8', cwd=BASE)
    banner = json.loads(_console_proc.stdout.readline())
    return banner


@app.post('/api/console')
async def console_exec(request: Request):
    """Run Python in the persistent console worker. Body: {"code": "..."}.
    Blocks until the code finishes (notebook semantics); use /api/console/restart
    to kill a runaway command."""
    if not _console_localonly(request):
        return JSONResponse({'error': 'console is localhost-only'}, status_code=403)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    code = (payload or {}).get('code', '')
    if not isinstance(code, str) or not code.strip():
        return {'out': '', 'err': None}
    def _exchange():
        global _console_proc
        with _console_lock:
            if _console_proc is None or _console_proc.poll() is not None:
                banner = _console_start()
                prefix = banner.get('out', '') + '\n'
            else:
                prefix = ''
            _console_proc.stdin.write(json.dumps({'code': code}) + '\n')
            _console_proc.stdin.flush()
            reply = json.loads(_console_proc.stdout.readline())
        reply['out'] = prefix + reply.get('out', '')
        return reply

    try:
        # off the event loop: a long-running command must not freeze the app
        return await asyncio.to_thread(_exchange)
    except Exception as e:
        return JSONResponse({'error': 'console worker failed: %s' % e},
                            status_code=500)


@app.post('/api/console/restart')
async def console_restart(request: Request):
    """Kill the worker (unsticks runaway code, clears all variables)."""
    if not _console_localonly(request):
        return JSONResponse({'error': 'console is localhost-only'}, status_code=403)
    def _do_restart():
        global _console_proc
        with _console_lock:
            if _console_proc is not None and _console_proc.poll() is None:
                _console_proc.kill()
            _console_proc = None
            return _console_start()

    try:
        banner = await asyncio.to_thread(_do_restart)
    except Exception as e:
        return JSONResponse({'error': 'restart failed: %s' % e}, status_code=500)
    return {'restarted': True, 'out': banner.get('out', '')}


def _read_positions_by_symbol(run_dir):
    """Group a run's archived position_log.csv by symbol (the full detail the
    Positions table shows). Empty dict if the file is missing."""
    import csv as _csv
    path = os.path.join(run_dir, 'position_log.csv')
    out = {}
    if not os.path.exists(path):
        return out
    numeric = {'entry_price', 'exit_price', 'size', 'pnl', 'pnlcomm'}
    with open(path, newline='', encoding='utf-8') as f:
        for row in _csv.DictReader(f):
            d = {}
            for k, v in row.items():
                if k in numeric and v not in (None, ''):
                    try:
                        d[k] = float(v)
                    except ValueError:
                        d[k] = v
                elif k == 'bars_held' and v not in (None, ''):
                    try:
                        d[k] = int(float(v))
                    except ValueError:
                        d[k] = v
                else:
                    d[k] = v
            out.setdefault(row.get('symbol', '?'), []).append(d)
    return out


@app.get('/api/testruns')
def list_test_runs(include_jobs: int = 0):
    """Archived backtest folders (reports/test_data/<tag>/) that carry a
    run.json snapshot — the ones the dashboard can reload / the report can
    compare. Newest first. include_jobs=1 prepends finished JOBS (tag
    'job:<id>') whose folders still exist — those carry chart data, so the
    comparison report can draw per-symbol prediction plots for them."""
    out = []
    if include_jobs:
        with _jobs_lock:
            done = [m for m in _jobs.values() if m['state'] == 'done']
        for m in sorted(done, key=lambda x: x['created'], reverse=True):
            if not os.path.isfile(os.path.join(JOBS_DIR, m['id'], 'results.json')):
                continue
            s = m.get('summary') or {}
            out.append({'tag': 'job:' + m['id'], 'kind': 'job',
                        'generated': m.get('created'),
                        'strategy': m['strategy'], 'timeframe': m['timeframe'],
                        'pnl': s.get('pnl'), 'return_pct': s.get('return_pct'),
                        'trades': s.get('trades_closed'), 'symbols': None,
                        'has_charts': os.path.isdir(os.path.join(JOBS_DIR, m['id'], 'chartdata'))})
    tdir = run_backtest.TEST_DATA_DIR
    if os.path.isdir(tdir):
        for tag in sorted(os.listdir(tdir), reverse=True):
            snap = _read_json(os.path.join(tdir, tag, 'run.json'))
            if not snap:
                continue
            p, s = snap.get('params', {}), snap.get('summary', {})
            out.append({'tag': tag, 'kind': 'archive', 'generated': snap.get('generated'),
                        'strategy': p.get('strategy'), 'timeframe': p.get('timeframe', '1m'),
                        'pnl': s.get('pnl'), 'return_pct': s.get('return_pct'),
                        'trades': s.get('trades_closed'), 'symbols': s.get('symbols'),
                        'window_start': p.get('window_start'), 'window_end': p.get('window_end')})
    return {'runs': out}


@app.get('/api/testrun/{tag}')
def load_test_run(tag: str):
    """Load one archived run for the dashboard: its run.json snapshot plus the
    per-symbol positions from position_log.csv. No candle/indicator chart data
    (not stored) — the UI shows PnL + positions only for loaded runs."""
    if not tag.replace('-', '').isalnum():           # tag is a timestamp folder
        return JSONResponse({'error': 'bad tag'}, status_code=400)
    run_dir = os.path.join(run_backtest.TEST_DATA_DIR, tag)
    snap = _read_json(os.path.join(run_dir, 'run.json'))
    if not snap:
        return JSONResponse({'error': 'no such archived run (needs run.json)'},
                            status_code=404)
    pos = _read_positions_by_symbol(run_dir)
    snap['charts'] = []                              # no candles for loaded runs
    snap['positionsBySymbol'] = pos
    snap['loaded_from'] = tag
    return snap


# generated-timestamp -> folder caches (run.json / results.json are immutable,
# so cache entries never go stale; missing folders just drop out of the map)
_gen2archive: dict = {}   # archive tag -> generated
_gen2job: dict = {}       # job id -> generated


def _gen_key(iso):
    """Timezone-proof matching key for a 'generated' timestamp: postgres
    returns it localized (+05:30) while run.json/results.json store UTC.
    Full microsecond precision — parallel jobs can finish mere microseconds
    apart, and a coarser key (this used to round to ms) made two such runs
    collide onto ONE links entry, cross-linking a History row to the wrong
    job folder (wrong name/charts, duplicated runs in compare reports)."""
    if not iso:
        return None
    try:
        return round(datetime.fromisoformat(str(iso)).timestamp(), 6)
    except (TypeError, ValueError):
        return str(iso)


_job2tag: dict = {}       # job id -> run_tag (from its results.json)


def _folder_links():
    """Two link maps so history rows (postgres) can be tied back to their
    artifact folders: by run_tag (exact — new rows store it) and by
    generated-timestamp key (fallback for rows predating the run_tag
    column). The archive folder's NAME is the run_tag itself."""
    links = {}       # gen_key -> {'archive_tag':…, 'job_id':…}
    tag_links = {}   # run_tag -> {'archive_tag':…, 'job_id':…}
    tdir = run_backtest.TEST_DATA_DIR
    for tag in (os.listdir(tdir) if os.path.isdir(tdir) else []):
        if not os.path.isdir(os.path.join(tdir, tag)):
            continue
        tag_links.setdefault(tag, {})['archive_tag'] = tag
        if tag not in _gen2archive:
            snap = _read_json(os.path.join(tdir, tag, 'run.json'))
            _gen2archive[tag] = _gen_key((snap or {}).get('generated'))
        g = _gen2archive[tag]
        if g:
            links.setdefault(g, {})['archive_tag'] = tag
    with _jobs_lock:
        done = [m['id'] for m in _jobs.values() if m['state'] == 'done']
    for jid in done:
        if jid not in _gen2job or jid not in _job2tag:
            res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
            _gen2job[jid] = _gen_key((res or {}).get('generated'))
            _job2tag[jid] = (res or {}).get('run_tag')
        if not os.path.isfile(os.path.join(JOBS_DIR, jid, 'results.json')):
            continue
        if _job2tag.get(jid):
            tag_links.setdefault(_job2tag[jid], {})['job_id'] = jid
        if _gen2job[jid]:
            links.setdefault(_gen2job[jid], {})['job_id'] = jid
    return links, tag_links


@app.get('/api/symbol-classes')
def symbol_classes_api():
    """{symbol: class} tags (Layer 1, DeFi, Meme, AI, ...) for every symbol
    this project trades — mirrors Binance's small faded category tag."""
    return {'classes': symbol_classes.SYMBOL_CLASS}


@app.get('/api/runs')
def runs_history(limit: int = 500):
    """Past backtest runs from Postgres (falls back cleanly when DB is down),
    enriched with links to their artifact folders: `job_id` (chart data,
    per-symbol positions) and `archive_tag` (test_data: position log, xlsx)."""
    try:
        import db
        conn = db.get_conn()
        out = db.list_runs(conn, limit=max(1, min(int(limit), 5000)))
        conn.close()
    except Exception as e:
        return JSONResponse({'error': 'database unavailable: %s' % e},
                            status_code=503)
    links, tag_links = _folder_links()
    for r in out:
        # exact run_tag link when the row has one (all new rows store it);
        # timestamp fallback only for legacy rows without it
        rt = r.get('run_tag')
        if rt:
            r.update(tag_links.get(rt, {}))
        else:
            r.update(links.get(_gen_key(r.get('generated')), {}))
        jid, tag = r.get('job_id'), r.get('archive_tag')
        artifact_dirs = []
        if jid:
            artifact_dirs.append(os.path.join(JOBS_DIR, jid))
        if tag:
            artifact_dirs.append(os.path.join(run_backtest.TEST_DATA_DIR, tag))
        r['has_crosssectional_xlsx'] = any(os.path.isfile(os.path.join(d, 'crosssectionaltests.xlsx'))
                                          for d in artifact_dirs)
        r['has_portfolio_xlsx'] = any(os.path.isfile(os.path.join(d, 'portfolio_stats.xlsx')) or
                                      os.path.isfile(os.path.join(d, 'report.xlsx'))
                                      for d in artifact_dirs)
        # friendly run name — not stored in postgres, read from whichever
        # artifact folder still exists (falls back to the run tag/job id,
        # same default the run itself uses when no name was given)
        r['name'] = None
        if jid:
            snap = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
            r['name'] = (snap or {}).get('name')
        if not r['name'] and tag:
            snap = _read_json(os.path.join(run_backtest.TEST_DATA_DIR, tag, 'run.json'))
            r['name'] = (snap or {}).get('name')
        r['name'] = r['name'] or jid or tag
    return {'runs': out}


@app.post('/api/testdata/delete')
async def delete_testdata(request: Request):
    """Delete ONE archived run folder (reports/test_data/<tag>)."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    tag = str(payload.get('tag') or '')
    if not tag or not all(c.isalnum() or c in '-_' for c in tag):
        return JSONResponse({'error': 'bad tag'}, status_code=400)
    p = os.path.join(run_backtest.TEST_DATA_DIR, tag)
    if not os.path.isdir(p):
        return JSONResponse({'error': 'no such archive'}, status_code=404)
    shutil.rmtree(p, ignore_errors=True)
    return {'ok': True}


@app.post('/api/runs/delete')
async def delete_run_row(request: Request):
    """Delete ONE run row (and its trades/equity/per_symbol) from postgres.
    Does not touch reports/test_data/<tag> — pair with /api/testdata/delete
    to remove both."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        run_id = int(payload.get('id'))
    except (TypeError, ValueError):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    import db
    conn = db.get_conn()
    try:
        ok = db.delete_run(conn, run_id)
    finally:
        conn.close()
    if not ok:
        return JSONResponse({'error': 'no such run'}, status_code=404)
    return {'ok': True}


@app.post('/api/sync')
async def sync_data(request: Request):
    """Incremental market-data sync — the exact routine launch.bat runs
    (fetch_binance_csv.main): asks postgres what's missing over the last N
    days, downloads only the gaps, upserts. Poll /api/sync/status."""
    global _sync_thread
    if _sync_thread is not None and _sync_thread.is_alive():
        return JSONResponse({'error': 'sync already running'}, status_code=409)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        days = max(1, min(int((payload or {}).get('days', 180)), 2000))
    except (TypeError, ValueError):
        days = 180
    interval = (payload or {}).get('interval')
    interval = interval if interval in ('1m', '1h') else '1m'

    def _work():
        import time as _time
        _sync_state.update(state='syncing', days=days, interval=interval,
                           started=_time.time(), error=None)
        try:
            import fetch_binance_csv
            fetch_binance_csv = importlib.reload(fetch_binance_csv)
            stats = fetch_binance_csv.main(days, interval) or {}
            _sync_state.update(state='done', **stats)
        except Exception as e:
            _sync_state.update(state='error', error=str(e))

    _sync_thread = threading.Thread(target=_work, daemon=True)
    _sync_thread.start()
    return {'started': True, 'days': days}


@app.get('/api/sync/status')
def sync_status():
    st = dict(_sync_state)
    st['thread_alive'] = _sync_thread is not None and _sync_thread.is_alive()
    if st.get('state') == 'syncing' and not st['thread_alive']:
        st['state'] = 'error'
        st.setdefault('error', 'sync thread stopped unexpectedly')
    return st


# ---- saved per-strategy / per-timeframe parameter configs ----------------
CONFIG_STORE = os.path.join(BASE, 'saved_configs.json')
_config_lock = threading.Lock()


def _load_configs():
    try:
        with open(CONFIG_STORE, encoding='utf-8') as f:
            d = json.load(f)
        return d if isinstance(d, list) else []
    except (OSError, ValueError):
        return []


def _write_configs(items):
    with open(CONFIG_STORE, 'w', encoding='utf-8') as f:
        json.dump(items, f, indent=1)


@app.get('/api/configs')
def list_configs():
    """All saved configs: [{id, strategy, timeframe, name, params, notes,
    default, saved}]. `default` is per (strategy, timeframe)."""
    with _config_lock:
        return {'configs': _load_configs()}


@app.post('/api/configs')
async def modify_configs(request: Request):
    """action=save {strategy,timeframe,name,params,notes} | default {id} |
    notes {id,notes} | delete {id}."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    action = payload.get('action')
    with _config_lock:
        items = _load_configs()
        if action == 'save':
            strategy = str(payload.get('strategy') or '')
            timeframe = payload.get('timeframe') if payload.get('timeframe') in ('1m', '1h') else '1m'
            params = payload.get('params')
            if not strategy or not isinstance(params, dict):
                return JSONResponse({'error': 'strategy and params required'}, status_code=400)
            cid = datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S-') + strategy
            first_for_pair = not any(c['strategy'] == strategy and c['timeframe'] == timeframe
                                     for c in items)
            items.append({'id': cid, 'strategy': strategy, 'timeframe': timeframe,
                          'name': str(payload.get('name') or '')[:80] or cid,
                          'params': params,
                          'notes': str(payload.get('notes') or '')[:2000],
                          'default': first_for_pair,   # first save = default
                          'saved': datetime.now(timezone.utc).isoformat(timespec='seconds')})
        elif action == 'default':
            cid = payload.get('id')
            hit = next((c for c in items if c['id'] == cid), None)
            if hit is None:
                return JSONResponse({'error': 'unknown config id'}, status_code=404)
            for c in items:
                if c['strategy'] == hit['strategy'] and c['timeframe'] == hit['timeframe']:
                    c['default'] = (c['id'] == cid)
        elif action == 'notes':
            hit = next((c for c in items if c['id'] == payload.get('id')), None)
            if hit is None:
                return JSONResponse({'error': 'unknown config id'}, status_code=404)
            hit['notes'] = str(payload.get('notes') or '')[:2000]
        elif action == 'rename':
            hit = next((c for c in items if c['id'] == payload.get('id')), None)
            if hit is None:
                return JSONResponse({'error': 'unknown config id'}, status_code=404)
            name = str(payload.get('name') or '').strip()[:80]
            if not name:
                return JSONResponse({'error': 'name required'}, status_code=400)
            hit['name'] = name
        elif action == 'delete':
            items = [c for c in items if c['id'] != payload.get('id')]
        else:
            return JSONResponse({'error': 'unknown action'}, status_code=400)
        _write_configs(items)
        return {'ok': True, 'configs': items}


# ---- parallel backtest jobs ------------------------------------------------
# Each job runs in its OWN process (backend/job_runner.py) writing artifacts
# to reports/jobs/<id>/. The manager thread starts queued jobs up to
# MAX_PARALLEL_JOBS, mirrors their status files into _jobs, and persists the
# registry so the list survives server restarts.
import itertools
import subprocess
import time as _t

JOBS_DIR = os.path.join(REPORTS, 'jobs')
JOBS_INDEX = os.path.join(JOBS_DIR, 'index.json')
MAX_PARALLEL_JOBS = int(os.getenv('MAX_PARALLEL_JOBS', '2'))
MAX_JOBS_PER_SUBMIT = 64

_jobs: dict = {}          # id -> meta dict
_job_procs: dict = {}     # id -> subprocess.Popen
_job_logs: dict = {}      # id -> open run.log file handle (closed once the job finishes)
_jobs_lock = threading.Lock()
_job_seq = 0


def _jobs_save():
    os.makedirs(JOBS_DIR, exist_ok=True)
    with open(JOBS_INDEX, 'w', encoding='utf-8') as f:
        json.dump(list(_jobs.values()), f, indent=1)


def _jobs_load():
    try:
        with open(JOBS_INDEX, encoding='utf-8') as f:
            for m in json.load(f):
                if m.get('state') in ('queued', 'running'):
                    m['state'] = 'error'
                    m['error'] = 'server restarted mid-job'
                _jobs[m['id']] = m
    except (OSError, ValueError):
        pass


def _expand_value(v):
    """One param value -> list of values. Strings may be 'start:stop:step'
    (inclusive) or 'a,b,c'; anything else is a scalar."""
    if not isinstance(v, str):
        return [v]
    s = v.strip()
    if not s:
        return [None]                     # empty -> use the strategy default
    if ':' in s:
        try:
            nums = [float(p) for p in s.split(':')]
        except ValueError:
            return [s]
        start, stop = nums[0], nums[1]
        step = nums[2] if len(nums) > 2 else 1.0
        if step == 0:
            return [start]
        out, x = [], start
        for _ in range(500):
            if (step > 0 and x > stop + 1e-9) or (step < 0 and x < stop - 1e-9):
                break
            out.append(int(x) if float(x).is_integer() else round(x, 10))
            x += step
        return out or [start]
    if ',' in s:
        out = []
        for p in (q.strip() for q in s.split(',')):
            if not p:
                continue
            try:
                f = float(p)
                out.append(int(f) if f.is_integer() else f)
            except ValueError:
                out.append(True if p.lower() == 'true'
                           else False if p.lower() == 'false' else p)
        return out or [None]
    try:
        f = float(s)
        return [int(f) if f.is_integer() else f]
    except ValueError:
        return [True if s.lower() == 'true'
                else False if s.lower() == 'false' else s]


import re as _re
_BRACKET_RE = _re.compile(r'\[([^\]]*)\]')


def _expand_brackets(v):
    """One param value -> list of BRACKET GROUPS, each a list of concrete
    values. A plain value (no brackets) is a single group, e.g. '2,3' ->
    [[2,3]]. Bracketed values split into matched groups: '[1],[2,3]' ->
    [[1], [2,3]]. Within a bracket the usual range/list syntax applies."""
    if isinstance(v, str) and '[' in v:
        groups = _BRACKET_RE.findall(v)
        if groups:
            return [_expand_value(g) for g in groups]
    return [_expand_value(v)]


def _expand_spec(spec):
    """One submission entry -> list of concrete param dicts.

    Bracket groups are MATCHED positionally across params (bracket 1 of param A
    with bracket 1 of param B, …); within a matched position the params form a
    cartesian PRODUCT; the per-position products are then unioned. Params with a
    single group (no brackets) broadcast to every position — so plain ranges/
    lists keep the old full-cartesian behaviour. Every param that uses more than
    one bracket must use the SAME number of brackets, else ValueError.

    e.g.  k=[1],[2,3]  a=[2],[3]  ->  (k1,a2), (k2,a3), (k3,a3)."""
    params = spec.get('params') or {}
    per = {k: _expand_brackets(v) for k, v in params.items()}
    multi = [len(g) for g in per.values() if len(g) > 1]
    n = max(multi) if multi else 1
    for k, g in per.items():
        if len(g) not in (1, n):
            raise ValueError(
                'param %r has %d bracket groups but the matched count is %d — '
                'every bracketed param must use the same number of [] groups'
                % (k, len(g), n))
    out, seen = [], set()
    for i in range(n):
        keys, valsets = [], []
        for k, g in per.items():
            keys.append(k)
            valsets.append(g[i] if len(g) == n else g[0])   # broadcast singletons
        for combo in itertools.product(*valsets) if keys else [()]:
            p = {k: v for k, v in zip(keys, combo) if v is not None}
            sig = tuple(sorted((k, repr(v)) for k, v in p.items()))
            if sig in seen:
                continue
            seen.add(sig)
            out.append(p)
    return out


def _new_job(strategy, timeframe, days, params, name=None):
    global _job_seq
    _job_seq += 1
    jid = 'j%s-%02d' % (datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S'), _job_seq % 100)
    jdir = os.path.join(JOBS_DIR, jid)
    os.makedirs(os.path.join(jdir, 'chartdata'), exist_ok=True)
    job = {'strategy': strategy, 'timeframe': timeframe, 'days': days,
           'params': params, 'name': name}
    with open(os.path.join(jdir, 'job.json'), 'w', encoding='utf-8') as f:
        json.dump(job, f)
    meta = dict(job, id=jid, state='queued', phase='queued', dismissed=False,
                created=datetime.now(timezone.utc).isoformat(timespec='seconds'),
                summary=None, error=None, elapsed=None)
    _jobs[jid] = meta
    return meta


def _spawn_job(meta):
    jid = meta['id']
    jdir = os.path.join(JOBS_DIR, jid)
    env = dict(os.environ, JOB_SKIP_REFRESH='1')
    # run_backtest already prints the full traceback on failure — capture it
    # to a per-job log instead of DEVNULL, so a job.error like "'X'" (just
    # str(exception)) can be followed up with the real stack trace instead of
    # requiring a source read to guess at the cause.
    log_f = open(os.path.join(jdir, 'run.log'), 'w', encoding='utf-8')
    proc = subprocess.Popen(
        [sys.executable, os.path.join(_paths.BACKEND, 'job_runner.py'), jdir],
        cwd=_paths.ROOT, env=env,
        stdout=log_f, stderr=subprocess.STDOUT)
    _job_procs[jid] = proc
    _job_logs[jid] = log_f
    meta.update(state='running', phase='starting',
                started=datetime.now(timezone.utc).isoformat(timespec='seconds'))


def _refresh_csvs_for(timeframe):
    """Serialize CSV-cache refreshes in the parent; skip when a sibling job of
    the same timeframe is mid-run (its feeds must not be rewritten). The cache
    window is the WIDEST window any pending/running job of this timeframe
    needs — not the whole history."""
    if any(m['state'] == 'running' and m['timeframe'] == timeframe
           for m in _jobs.values()):
        return
    days = [m.get('days') or run_backtest.BACKTEST_DAYS
            for m in _jobs.values()
            if m['timeframe'] == timeframe and m['state'] in ('queued', 'running')]
    from datetime import timedelta
    need_start = (datetime.now(timezone.utc)
                  - timedelta(days=(max(days) if days else run_backtest.BACKTEST_DAYS) + 2))
    try:
        conn = db.get_conn()
        try:
            run_backtest._refresh_full_csvs(conn, timeframe, need_start=need_start)
        finally:
            conn.close()
    except Exception as e:
        print('[jobs] csv refresh failed (%s) — jobs will use existing CSVs' % e)


def _jobs_tick():
    with _jobs_lock:
        changed = False
        # reap running jobs
        for jid, proc in list(_job_procs.items()):
            meta = _jobs.get(jid)
            if meta is None:
                proc.kill(); _job_procs.pop(jid, None); continue
            st = _read_json(os.path.join(JOBS_DIR, jid, 'status.json')) or {}
            phase = st.get('phase') or st.get('state') or meta['phase']
            prog = ''
            if st.get('done') is not None and st.get('total'):
                prog = ' %s/%s' % (st['done'], st['total'])
            newphase = str(phase) + prog
            if newphase != meta['phase']:
                meta['phase'] = newphase; changed = True
            if proc.poll() is not None:                    # process ended
                _job_procs.pop(jid, None)
                log_f = _job_logs.pop(jid, None)
                if log_f:
                    try: log_f.close()
                    except OSError: pass
                res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
                if st.get('state') == 'done' and res:
                    # ~60-point equity sparkline for the job card
                    eq = [v for _dt, v in (res.get('equity') or [])]
                    step = max(1, len(eq) // 60)
                    meta.update(state='done', phase='done',
                                elapsed=st.get('elapsed'),
                                spark=[round(v, 2) for v in eq[::step]][:60],
                                summary={k: res.get('summary', {}).get(k) for k in
                                         ('pnl', 'return_pct', 'trades_closed',
                                          'trades_total', 'unrealised_pnl',
                                          'open_positions', 'win_rate_pct',
                                          'sharpe_arithmetic', 'max_drawdown_pct')})
                else:
                    meta.update(state='error', phase='error',
                                error=st.get('error') or 'process exited unexpectedly')
                changed = True
        # start queued jobs up to the parallel limit
        running = sum(1 for m in _jobs.values() if m['state'] == 'running')
        for meta in [m for m in _jobs.values() if m['state'] == 'queued']:
            if running >= MAX_PARALLEL_JOBS:
                break
            _refresh_csvs_for(meta['timeframe'])
            _spawn_job(meta)
            running += 1; changed = True
        if changed:
            _jobs_save()


def _jobs_manager():
    while True:
        try:
            _jobs_tick()
        except Exception as e:
            print('[jobs] manager error: %s' % e)
        _t.sleep(1.0)


_jobs_load()
threading.Thread(target=_jobs_manager, daemon=True).start()


def _job_artifact_flags(jid, state):
    """Cheap isfile checks — only for finished jobs, so this stays fast even
    at the 3s poll interval."""
    if state != 'done':
        return {}
    d = os.path.join(JOBS_DIR, jid)
    out = {
        'has_portfolio_xlsx': os.path.isfile(os.path.join(d, 'portfolio_stats.xlsx')),
        'has_crosssectional_xlsx': os.path.isfile(os.path.join(d, 'crosssectionaltests.xlsx')),
    }
    # the job's own 'name' (job.json) is whatever the user typed at submit
    # time, which is None when left blank — once done, prefer the run's
    # resolved name (name or run_tag) from results.json
    res = _read_json(os.path.join(d, 'results.json'))
    if res and res.get('name'):
        out['name'] = res['name']
    # summary keys added after a job's card summary was mirrored (trades_total,
    # unrealised_pnl, ...) won't be in the persisted meta for OLD done jobs —
    # backfill them here from results.json so the card shows them without a
    # re-run. list_jobs merges this over the meta summary.
    if res:
        full = res.get('summary') or {}
        out['_summary_extra'] = {k: full.get(k) for k in
                                 ('trades_total', 'unrealised_pnl', 'open_positions')
                                 if k in full}
        # the resolved backtest window (not in the submitted job params) — the
        # Compare tab shows it as a formatted period row
        rp = res.get('params') or {}
        for k in ('window_start', 'window_end'):
            if rp.get(k):
                out[k] = rp[k]
    return out


@app.get('/api/jobs')
def list_jobs():
    with _jobs_lock:
        jobs = []
        for m in _jobs.values():
            if m.get('dismissed'):
                continue
            flags = _job_artifact_flags(m['id'], m['state'])
            extra = flags.pop('_summary_extra', None)
            job = dict(m, **flags)
            if extra:
                job['summary'] = dict(m.get('summary') or {}, **extra)
            jobs.append(job)
        return {'jobs': jobs, 'max_parallel': MAX_PARALLEL_JOBS}


@app.post('/api/jobs')
async def submit_jobs(request: Request):
    """{jobs: [{strategy, timeframe, days, params}...]}. Param values may be
    scalars, 'start:stop:step' ranges or 'a,b,c' lists — every combination
    becomes its own job (cartesian per entry, capped)."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    specs = payload.get('jobs') or []
    if not isinstance(specs, list) or not specs:
        return JSONResponse({'error': 'no jobs given'}, status_code=400)
    expanded = []
    for spec in specs:
        name = spec.get('strategy')
        if name not in run_backtest.STRATEGIES:
            return JSONResponse({'error': 'unknown strategy %r' % name}, status_code=400)
        tf = spec.get('timeframe') if spec.get('timeframe') in ('1m', '1h') else '1m'
        try:
            days = int(spec.get('days') or 0) or None
        except (TypeError, ValueError):
            days = None
        # optional friendly label, doesn't need to be unique — falls back to
        # the run tag (the archive folder name) if left blank
        run_name = str(spec.get('name')).strip() if spec.get('name') else None
        try:
            combos = _expand_spec(spec)
        except ValueError as e:
            return JSONResponse({'error': str(e)}, status_code=400)
        for params in combos:
            expanded.append((name, tf, days, params, run_name))
    if len(expanded) > MAX_JOBS_PER_SUBMIT:
        return JSONResponse({'error': 'that expands to %d jobs (cap %d) — narrow the ranges'
                             % (len(expanded), MAX_JOBS_PER_SUBMIT)}, status_code=400)
    with _jobs_lock:
        created = [_new_job(*e)['id'] for e in expanded]
        _jobs_save()
    return {'created': created, 'count': len(created)}


@app.post('/api/jobs/action')
async def job_action(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    jid, action = payload.get('id'), payload.get('action')
    with _jobs_lock:
        meta = _jobs.get(jid)
        if meta is None:
            return JSONResponse({'error': 'unknown job'}, status_code=404)
        if action == 'cancel':
            proc = _job_procs.pop(jid, None)
            if proc is not None:
                proc.kill()
            log_f = _job_logs.pop(jid, None)
            if log_f:
                try: log_f.close()
                except OSError: pass
            if meta['state'] in ('queued', 'running'):
                meta.update(state='cancelled', phase='cancelled')
        elif action == 'dismiss':
            if meta['state'] in ('queued', 'running'):
                return JSONResponse({'error': 'cancel it first'}, status_code=400)
            meta['dismissed'] = True
        elif action == 'delete':
            # remove the job's FOLDER (chartdata, per-symbol positions, results)
            # and its registry entry — gone for good, unlike dismiss.
            if meta['state'] in ('queued', 'running'):
                return JSONResponse({'error': 'cancel it first'}, status_code=400)
            _cache_gen_key(jid)
            shutil.rmtree(os.path.join(JOBS_DIR, jid), ignore_errors=True)
            _jobs.pop(jid, None)
        else:
            return JSONResponse({'error': 'unknown action'}, status_code=400)
        _jobs_save()
    return {'ok': True}


def _write_json(path, obj):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(obj, f, indent=1, default=str)


@app.post('/api/rename')
async def rename_run(request: Request):
    """Rename a backtest run. Pass job_id and/or archive_tag (a run can have
    both, or an archive-only History row after its job folder was deleted) —
    every place that stores the friendly name gets patched so it stays
    consistent across the job card, History, and Compare."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    name = str(payload.get('name') or '').strip()[:200]
    if not name:
        return JSONResponse({'error': 'name required'}, status_code=400)
    jid = payload.get('job_id')
    tag = payload.get('archive_tag')
    if not jid and not tag:
        return JSONResponse({'error': 'job_id or archive_tag required'}, status_code=400)
    touched = False
    if jid:
        with _jobs_lock:
            meta = _jobs.get(jid)
            if meta is not None:
                meta['name'] = name
                _jobs_save()
                touched = True
        rp = os.path.join(JOBS_DIR, jid, 'results.json')
        res = _read_json(rp)
        if res is not None:
            res['name'] = name
            _write_json(rp, res)
            touched = True
    if tag:
        rp = os.path.join(run_backtest.TEST_DATA_DIR, tag, 'run.json')
        res = _read_json(rp)
        if res is not None:
            res['name'] = name
            _write_json(rp, res)
            touched = True
    if not touched:
        return JSONResponse({'error': 'no such job/archive'}, status_code=404)
    return {'ok': True, 'name': name}


@app.post('/api/history/clear')
async def clear_history(request: Request):
    """what='tests': wipe reports/test_data archives + the postgres run
    history. what='jobs': delete every finished job folder + registry entry
    (running/queued jobs are left alone)."""
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    what = payload.get('what')
    if what == 'tests':
        n = 0
        tdir = run_backtest.TEST_DATA_DIR
        for name in (os.listdir(tdir) if os.path.isdir(tdir) else []):
            p = os.path.join(tdir, name)
            if os.path.isdir(p):
                shutil.rmtree(p, ignore_errors=True)
                n += 1
        try:
            conn = db.get_conn()
            try:
                db.clear_run_history(conn)
            finally:
                conn.close()
        except Exception as e:
            return JSONResponse({'error': 'archives deleted (%d) but DB clear failed: %s'
                                 % (n, e)}, status_code=500)
        return {'ok': True, 'deleted_archives': n, 'db_cleared': True}
    if what == 'jobs':
        n = 0
        with _jobs_lock:
            for jid, meta in list(_jobs.items()):
                if meta['state'] in ('queued', 'running'):
                    continue
                _cache_gen_key(jid)
                shutil.rmtree(os.path.join(JOBS_DIR, jid), ignore_errors=True)
                _jobs.pop(jid, None)
                n += 1
            _jobs_save()
        return {'ok': True, 'deleted_jobs': n}
    return JSONResponse({'error': "what must be 'tests' or 'jobs'"}, status_code=400)


@app.get('/api/jobs/{jid}/results')
def job_results(jid: str):
    if not all(c.isalnum() or c in '-_' for c in jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
    if res is None:
        return JSONResponse({'error': 'no results for this job (yet)'}, status_code=404)
    res['chartbase'] = '/reports/jobs/%s/chartdata' % jid
    res['job_id'] = jid
    return res


@app.get('/api/jobs/{jid}/log')
def job_log(jid: str):
    """Full stdout/stderr of the job subprocess (includes the traceback
    behind a failed job's short error message). Useful when a job's
    'error' field is just str(exception) — e.g. a bare KeyError repr like
    "'X'" — and the actual cause needs the stack trace."""
    if not all(c.isalnum() or c in '-_' for c in jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    path = os.path.join(JOBS_DIR, jid, 'run.log')
    if not os.path.isfile(path):
        return JSONResponse({'error': 'no log for this job (predates run.log capture, '
                             'or the job is still running)'}, status_code=404)
    with open(path, encoding='utf-8', errors='replace') as f:
        return PlainTextResponse(f.read())


def _safe_job_id(jid):
    return bool(jid) and all(c.isalnum() or c in '-_' for c in str(jid))


def _cache_gen_key(jid):
    """Remember a job's run_tag + 'generated' key before its folder is
    deleted, so _archive_for_job can still find the matching test_data
    archive afterwards (results.json won't be readable anymore)."""
    if jid in _gen2job and jid in _job2tag:
        return
    res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json')) or {}
    key = _gen_key(res.get('generated'))
    if key:
        _gen2job[jid] = key
    if res.get('run_tag'):
        _job2tag[jid] = res['run_tag']


def _archive_for_job(jid):
    """Find the test_data archive belonging to a completed job, if present.
    The archive folder's name IS the run's run_tag, so that's an exact
    match; generated-timestamp scan is kept only as a fallback for old job
    folders whose results.json predates the run_tag field. Uses the
    _gen2job/_job2tag caches when the job folder has been deleted."""
    res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
    tdir = run_backtest.TEST_DATA_DIR
    run_tag = (res or {}).get('run_tag') or _job2tag.get(jid)
    if run_tag:
        p = os.path.join(tdir, run_tag)
        return p if os.path.isdir(p) else None
    key = _gen_key(res.get('generated')) if res else _gen2job.get(jid)
    if not key or not os.path.isdir(tdir):
        return None
    for tag in os.listdir(tdir):
        snap = _read_json(os.path.join(tdir, tag, 'run.json')) or {}
        if _gen_key(snap.get('generated')) == key:
            return os.path.join(tdir, tag)
    return None


def _artifact_paths(jid, filename):
    paths = []
    job_dir = os.path.join(JOBS_DIR, jid)
    if os.path.isdir(job_dir):
        paths.append(os.path.join(job_dir, filename))
    archive = _archive_for_job(jid)
    if archive:
        paths.append(os.path.join(archive, filename))
    return paths


def _save_crosssectional_xlsx(path, test_name, data):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        wb = load_workbook(path) if os.path.isfile(path) else Workbook()
    except Exception:
        wb = Workbook()
    if wb.sheetnames == ['Sheet'] and wb['Sheet'].max_row == 1 and wb['Sheet']['A1'].value is None:
        del wb['Sheet']
    title = ''.join(c for c in str(test_name or 'test') if c.isalnum() or c in ' _-')[:28] or 'test'
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(bold=True, color='FFFFFF')
    label_font = Font(bold=True, color='374151')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    pos_font = Font(color='1A7F37')   # green
    neg_font = Font(color='C42B1C')   # red

    def header_row(row_idx):
        for c in ws[row_idx]:
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal='center')

    def sign_font(v):
        return pos_font if isinstance(v, (int, float)) and v > 0 \
            else neg_font if isinstance(v, (int, float)) and v < 0 else None

    ws.append(['Cross-sectional test', test_name or 'test'])
    ws.append(['Job', data.get('job')])
    ws.append(['Signal', data.get('signal')])
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = label_font
    ws.append([])
    ws.append(['Horizon', 'Mean IC', 'IC std', 'IR', 't-stat', '% positive', 'Observations'])
    header_row(5)
    horizons = data.get('horizons') or {}
    row = 6
    for h, item in horizons.items():
        if item.get('error'):
            ws.append([h, item.get('error')])
            row += 1
            continue
        ws.append([h, item.get('mean'), item.get('std'), item.get('ir'), item.get('tstat'),
                   item.get('pos_share'), item.get('n')])
        for col, key in ((2, 'mean'), (3, 'std'), (4, 'ir'), (5, 'tstat')):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.0000'
            f = sign_font(item.get(key))
            if f:
                cell.font = f
        pct_cell = ws.cell(row=row, column=6)
        pct_cell.number_format = '0.00%'
        ws.cell(row=row, column=7).number_format = '0'
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1
    ws.append([])
    ws.append(['Rolling series'])
    ws.cell(row=ws.max_row, column=1).font = label_font
    ws.append(['Horizon', 'Time', 'Value'])
    header_row(ws.max_row)
    for h, item in horizons.items():
        for point in item.get('series') or []:
            ws.append([h, point[0], point[1]])
            r = ws.max_row
            v_cell = ws.cell(row=r, column=3)
            v_cell.number_format = '0.0000'
            f = sign_font(point[1])
            if f:
                v_cell.font = f
    ws.freeze_panes = 'A6'
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(42, max(12, max(
            len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, ws.max_row + 1)) + 2))
    wb.save(path)


def _save_portfolio_xlsx(path, results):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Portfolio stats'
    ws.append(['Metric', 'Value'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F2937')
    for k, v in (results.get('pyfolio') or {}).get('stats', {}).items():
        ws.append([k, v])
    ws2 = wb.create_sheet('Parameters')
    ws2.append(['Parameter', 'Value'])
    for k, v in (results.get('params') or {}).items():
        ws2.append([k, v])
    for sh in wb.worksheets:
        sh.column_dimensions['A'].width = 32
        sh.column_dimensions['B'].width = 28
    wb.save(path)


@app.post('/api/test/export/{jid}')
async def export_crosssectional(jid: str, request: Request):
    if not _safe_job_id(jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    data = await request.json()
    test_name = str(data.get('test') or 'test')
    paths = _artifact_paths(jid, 'crosssectionaltests.xlsx')
    if not paths:
        return JSONResponse({'error': 'job folder not found'}, status_code=404)
    try:
        for path in paths:
            _save_crosssectional_xlsx(path, test_name, data.get('data') or {})
    except Exception as e:
        return JSONResponse({'error': 'could not save workbook: %s' % e}, status_code=500)
    return {'ok': True, 'filename': 'crosssectionaltests.xlsx', 'paths': len(paths)}


@app.get('/api/download/crosssectional/{jid}')
def download_crosssectional(jid: str):
    if not _safe_job_id(jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    path = os.path.join(JOBS_DIR, jid, 'crosssectionaltests.xlsx')
    if not os.path.isfile(path):
        return JSONResponse({'error': 'cross-sectional workbook has not been exported'}, status_code=404)
    return FileResponse(path, filename='crosssectionaltests.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get('/api/download/portfolio/{jid}')
def download_portfolio(jid: str):
    if not _safe_job_id(jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    paths = _artifact_paths(jid, 'portfolio_stats.xlsx')
    job_path = paths[0] if paths else None
    res = _read_json(os.path.join(JOBS_DIR, jid, 'results.json'))
    if res is None or not job_path:
        return JSONResponse({'error': 'job results are unavailable'}, status_code=404)
    try:
        _save_portfolio_xlsx(job_path, res)
        for archive_path in paths[1:]:
            shutil.copy2(job_path, archive_path)
    except Exception as e:
        return JSONResponse({'error': 'could not save workbook: %s' % e}, status_code=500)
    return FileResponse(job_path, filename='portfolio_stats.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---- testing hub: cross-sectional analytics on a finished job -------------
_read_position_pnls = pnl_stats.read_position_pnls
_pnl_stats = pnl_stats.pnl_stats


def _read_unrealised(jid):
    """{symbol: unrealised_mtm} for positions open at the end of the run
    (written by run_backtest as unrealised.json). Empty when absent — runs
    predating the feature, or runs that ended flat."""
    return _read_json(os.path.join(JOBS_DIR, jid, 'unrealised.json')) or {}


@app.get('/api/test/contrib/{jid}')
def per_symbol_contribution(jid: str):
    """Per-symbol P&L attribution for one job, from its FULL (uncapped)
    per-symbol position logs in reports/jobs/<jid>/positions/<SYMBOL>.csv.
    Each symbol: closed-trade count, total net (realised) pnl, win rate, a
    trade-level Sharpe (mean/std of per-trade net pnl, ×√n — dimensionless,
    not annualized), the max drawdown of the running cumulative net pnl (in
    USDT), and `unrealised` = mark-to-market of any position still open at
    the end (realised + unrealised should reconcile to the run's total).
    Sorted by total realised pnl descending."""
    if not _safe_job_id(jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    pdir = os.path.join(JOBS_DIR, jid, 'positions')
    if not os.path.isdir(pdir):
        return JSONResponse({'error': 'job has no per-symbol position logs'},
                            status_code=404)
    by_sym = _read_position_pnls(pdir)
    unreal = _read_unrealised(jid)
    # a symbol can end holding an open position without ever closing one, so
    # union both key sets
    rows = []
    for sym in set(by_sym) | set(unreal):
        stats = _pnl_stats(by_sym[sym]) if sym in by_sym else _pnl_stats([])
        rows.append(dict(symbol=sym, unrealised=round(float(unreal.get(sym, 0.0)), 4), **stats))
    rows.sort(key=lambda r: r['pnl'], reverse=True)
    return {'job': jid, 'symbols': len(rows), 'rows': rows}


@app.get('/api/test/contrib-class/{jid}')
def per_class_contribution(jid: str):
    """Same P&L attribution as /api/test/contrib, but pooled by sector/
    category tag (symbol_classes.SYMBOL_CLASS) — every trade from every
    symbol in a class gets combined into one trade-level stats line, so
    win rate / Sharpe / max DD are computed correctly (not just averaged
    across symbols). `unrealised` sums the class's open-position MTM."""
    if not _safe_job_id(jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    pdir = os.path.join(JOBS_DIR, jid, 'positions')
    if not os.path.isdir(pdir):
        return JSONResponse({'error': 'job has no per-symbol position logs'},
                            status_code=404)
    by_sym = _read_position_pnls(pdir)
    unreal = _read_unrealised(jid)
    by_class = {}
    class_syms = {}
    class_unreal = {}
    for sym in set(by_sym) | set(unreal):
        cls = symbol_classes.classify(sym)
        by_class.setdefault(cls, []).extend(by_sym.get(sym, []))
        class_syms.setdefault(cls, []).append(sym)
        class_unreal[cls] = class_unreal.get(cls, 0.0) + float(unreal.get(sym, 0.0))
    rows = [dict(cls=cls, symbols=sorted(class_syms[cls]),
                 unrealised=round(class_unreal.get(cls, 0.0), 4), **_pnl_stats(pnls))
            for cls, pnls in by_class.items()]
    rows.sort(key=lambda r: r['pnl'], reverse=True)
    return {'job': jid, 'classes': len(rows), 'rows': rows}


@app.get('/api/test/ic/{jid}')
def information_coefficient(jid: str, horizons: str = '1,4,24'):
    """Cross-sectional Information Coefficient for one job.

    Signal (per symbol, per bar): (prediction - close) / close — the filter's
    expected relative move (its first MAIN chart line is taken as the
    prediction). Forward return over horizon h: close[t+h]/close[t] - 1.
    IC_t = Spearman rank correlation across symbols between signal_t and the
    h-bar forward return, computed at every bar with >= 10 symbols. Summary
    per horizon: mean IC, std, IR (=mean/std), t-stat (=IR*sqrt(n)), share of
    positive ICs, observations. NOTE: for h>1 consecutive ICs overlap, so the
    plain t-stat overstates significance — treat it as indicative.
    """
    if not all(c.isalnum() or c in '-_' for c in jid):
        return JSONResponse({'error': 'bad id'}, status_code=400)
    cdir = os.path.join(JOBS_DIR, jid, 'chartdata')
    if not os.path.isdir(cdir):
        return JSONResponse({'error': 'job has no chart data'}, status_code=404)
    try:
        hs = sorted({max(1, int(h)) for h in horizons.split(',') if h.strip()})[:6]
    except ValueError:
        return JSONResponse({'error': 'bad horizons'}, status_code=400)

    import numpy as np
    import pandas as pd

    closes, preds = {}, {}
    for p in os.listdir(cdir):
        if not p.endswith('.json'):
            continue
        cd = _read_json(os.path.join(cdir, p))
        if not cd:
            continue
        sym = p[:-5]
        candles = cd.get('candles') or []
        main = [L for L in (cd.get('lines') or []) if L.get('kind') != 'additional']
        if len(candles) < 30 or not main or len(main[0].get('points') or []) < 30:
            continue
        closes[sym] = pd.Series({c['time']: c['close'] for c in candles})
        preds[sym] = pd.Series({q['time']: q['value'] for q in main[0]['points']})
    if len(closes) < 10:
        return JSONResponse({'error': 'need >=10 symbols with prediction lines '
                             '(got %d)' % len(closes)}, status_code=400)

    C = pd.DataFrame(closes)          # bars x symbols (aligned on epoch time)
    P = pd.DataFrame(preds).reindex(C.index)
    S = (P - C) / C                   # the cross-sectional signal

    def rowwise_ic(sig, fwd):
        both = sig.notna() & fwd.notna()
        n = both.sum(axis=1)
        rs = sig.where(both).rank(axis=1)
        rf = fwd.where(both).rank(axis=1)
        rs = rs.sub(rs.mean(axis=1), axis=0)
        rf = rf.sub(rf.mean(axis=1), axis=0)
        cov = (rs * rf).sum(axis=1)
        den = np.sqrt((rs ** 2).sum(axis=1) * (rf ** 2).sum(axis=1))
        ic = cov / den.replace(0, np.nan)
        return ic.where(n >= 10).dropna()

    out = {}
    for h in hs:
        F = C.shift(-h) / C - 1.0
        ic = rowwise_ic(S, F)
        if ic.empty:
            out[str(h)] = {'error': 'no overlapping observations'}
            continue
        mean, std, n = float(ic.mean()), float(ic.std(ddof=1)), int(len(ic))
        step = max(1, n // 500)
        out[str(h)] = {
            'mean': round(mean, 5), 'std': round(std, 5),
            'ir': round(mean / std, 4) if std > 0 else None,
            'tstat': round(mean / std * np.sqrt(n), 2) if std > 0 else None,
            'pos_share': round(float((ic > 0).mean()), 4), 'n': n,
            'series': [[int(t), round(float(v), 5)]
                       for t, v in ic.iloc[::step].items()],
        }
    return {'job': jid, 'symbols': len(closes), 'bars': len(C),
            'signal': '(prediction - close) / close from each symbol\'s main line',
            'horizons': out}


@app.get('/api/logs')
def get_logs(since: int = 0, limit: int = 800):
    """Server log lines with id > `since` (tail `limit`). The dashboard's log
    window polls this incrementally."""
    with _log_lock:
        lines = [{'id': i, 't': t, 'line': l}
                 for (i, t, l) in _LOG_BUF if i > since]
    lines = lines[-max(1, min(limit, 2000)):]
    return {'lines': lines,
            'last': lines[-1]['id'] if lines else since}


@app.get('/api/data/status')
def data_status(interval: str = '1m', spike_pct: float = 0.5):
    """Full per-symbol data health for `interval`. Beyond span + gap count
    (missing bars over each symbol's own 24/7 span), this now also reports
    VALUE integrity — non-positive/null OHLC, high<low, zero/negative volume,
    suspicious single-bar jumps over `spike_pct` (fraction, e.g. 0.5 = 50%) —
    and the stored price range, so a corrupted symbol (the kind that can
    silently swing a whole backtest) is visible, not just gaps. Also returns
    portfolio-level totals for the header."""
    interval = interval if interval in ('1m', '1h') else '1m'
    bar_s = 3600 if interval == '1h' else 60
    try:
        spike_pct = max(0.01, min(float(spike_pct), 100.0))
    except (TypeError, ValueError):
        spike_pct = 0.5
    try:
        import db
        conn = db.get_conn()
    except Exception as e:
        return JSONResponse({'error': 'db unavailable: %s' % e}, status_code=503)
    try:
        info = db.data_integrity(conn, interval, spike_pct)
    finally:
        conn.close()
    now = datetime.now(timezone.utc)
    out = []
    for sym in sorted(info):
        d = info[sym]
        mn, mx, cnt = d['first'], d['last'], d['bars']
        expected = int((mx - mn).total_seconds() // bar_s) + 1
        missing = max(0, expected - cnt)
        # any value-level problem (gaps are tracked separately as `missing`)
        value_issues = (d['bad_ohlc'] + d['hl_viol'] + d['neg_vol'] + d['spikes'])
        out.append({
            'symbol': sym,
            'first': mn.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M'),
            'last': mx.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M'),
            'age_min': round((now - mx).total_seconds() / 60.0, 1),
            'bars': cnt,
            'expected': expected,
            'missing': missing,
            'completeness': round(100.0 * cnt / expected, 2) if expected else 100.0,
            'bad_ohlc': d['bad_ohlc'],
            'hl_viol': d['hl_viol'],
            'zero_vol': d['zero_vol'],
            'neg_vol': d['neg_vol'],
            'spikes': d['spikes'],
            'min_close': float(d['min_close']) if d['min_close'] is not None else None,
            'max_close': float(d['max_close']) if d['max_close'] is not None else None,
            'value_issues': value_issues,
            'ok': missing == 0 and value_issues == 0,
        })
    totals = {
        'symbols': len(out),
        'bars': sum(s['bars'] for s in out),
        'with_gaps': sum(1 for s in out if s['missing'] > 0),
        'with_value_issues': sum(1 for s in out if s['value_issues'] > 0),
        'clean': sum(1 for s in out if s['ok']),
        'stale': sum(1 for s in out if s['age_min'] and s['age_min'] > 1440),
        'earliest': min((s['first'] for s in out), default=None),
        'latest': max((s['last'] for s in out), default=None),
    }
    return {'interval': interval, 'generated': now.strftime('%Y-%m-%d %H:%M'),
            'totals': totals, 'symbols': out}


@app.get('/api/data/spikes/{symbol}')
def data_spikes(symbol: str, interval: str = '1m', spike_pct: float = 0.5):
    """The actual bars where |close/prev_close - 1| > spike_pct for one symbol
    — so a flagged symbol can be reviewed bar-by-bar (real move vs bad tick)."""
    if not symbol or not all(c.isalnum() for c in symbol):
        return JSONResponse({'error': 'bad symbol'}, status_code=400)
    interval = interval if interval in ('1m', '1h') else '1m'
    try:
        spike_pct = max(0.01, min(float(spike_pct), 100.0))
    except (TypeError, ValueError):
        spike_pct = 0.5
    try:
        import db
        conn = db.get_conn()
    except Exception as e:
        return JSONResponse({'error': 'db unavailable: %s' % e}, status_code=503)
    try:
        rows = db.spike_bars(conn, symbol, interval, spike_pct)
    finally:
        conn.close()
    out = []
    for ts, prev, close, high, low, vol in rows:
        prev = float(prev); close = float(close)
        out.append({
            'ts': ts.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M'),
            'prev_close': prev, 'close': close,
            'pct': round((close / prev - 1) * 100, 2),
            'high': float(high), 'low': float(low), 'volume': float(vol),
        })
    return {'symbol': symbol, 'interval': interval,
            'spike_pct': spike_pct, 'count': len(out), 'bars': out}


@app.get('/api/download/pyfolio')
def download_pyfolio():
    if not os.path.isdir(run_backtest.TEST_DATA_DIR):
        return JSONResponse({'error': 'no report yet'}, status_code=404)
    zip_base = os.path.join(REPORTS, 'test_data_report')
    shutil.make_archive(zip_base, 'zip', run_backtest.TEST_DATA_DIR)
    return FileResponse(zip_base + '.zip', filename='test_data_report.zip')


@app.on_event('startup')
def _migrate_schema():
    """Best-effort schema migration at boot (adds e.g. runs.run_tag) so
    /api/runs doesn't 500 on a pre-migration database before the first
    backtest gets a chance to run init_schema. DB down is fine — every
    query path already degrades gracefully."""
    try:
        conn = db.get_conn(_retries=1)
        try:
            db.init_schema(conn)
        finally:
            conn.close()
    except Exception as e:
        print('[server] schema migration skipped (%s)' % e)


if __name__ == '__main__':
    uvicorn.run(app, host='127.0.0.1', port=8001)
